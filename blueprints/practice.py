# ==============================================================================
# FILE: blueprints/practice.py
# Practice blueprint – MCQ, Fill-in-the-Blanks, Unscramble, Self-Study,
# public links, QR shortcuts, PDF export, score views, CSV/Excel export
# ==============================================================================

import json
import csv
import re
import os
import secrets
from io import BytesIO, StringIO
from datetime import datetime

from flask import (
    Blueprint, request, session, jsonify, redirect, url_for,
    render_template, abort, flash, Response,
)

from models import (
    Topic, PracticeQuestion, PracticeLink, PracticeSubmission,
    AttemptHistory, Classroom, ClassroomStudent, GameQuestion,
    UserSubscription,
)
from blueprints.helpers import login_required, _get_topic_or_404, _json_error

practice_bp = Blueprint("practice", __name__)


# ==============================================================================
# Helper Functions
# ==============================================================================
def _normalize_practice_questions(rows):
    out = []
    for r in rows:
        q = dict(r)
        prompt, choices = "", []
        raw = q.get("question") or ""
        try:
            obj = json.loads(raw)
            if isinstance(obj, dict):
                prompt = (obj.get("prompt") or "").strip()
                choices = [str(x) for x in (obj.get("choices") or [])]
            else:
                prompt = str(obj)
        except Exception:
            prompt = str(raw)
        out.append({
            "id": q.get("id"),
            "prompt": prompt,
            "choices": choices,
            "correct_answer": q.get("correct_answer") or "",
        })
    return out


def _build_practice_pdf(topic_title, questions, include_answers=False):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    from reportlab.lib.utils import simpleSplit

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    mx, y, lh, mw = 2 * cm, h - 2 * cm, 14, w - 4 * cm

    def draw(lines, y0):
        y = y0
        for ln in lines:
            if y < 2 * cm:
                c.showPage()
                y = h - 2 * cm
            c.drawString(mx, y, ln)
            y -= lh
        return y

    c.setFont("Helvetica-Bold", 16)
    y = draw([f"Practice: {topic_title}"], y)
    c.setFont("Helvetica", 11)
    y = draw(["Name: ________________________   Class: __________", ""], y)

    for i, q in enumerate(questions, 1):
        c.setFont("Helvetica-Bold", 12)
        y = draw(simpleSplit(f"{i}. {q.get('prompt', '')}", "Helvetica-Bold", 12, mw), y)
        c.setFont("Helvetica", 11)
        ch = q.get("choices") or []
        if len(ch) == 4:
            for lab, cv in zip(["A", "B", "C", "D"], ch):
                y = draw(simpleSplit(f"   ({lab}) {cv}", "Helvetica", 11, mw), y)
        if include_answers:
            y = draw([f"   Answer: {q.get('correct_answer', '')}"], y)
        y = draw([""], y)

    c.showPage()
    c.save()
    return buf.getvalue()


def _get_practice_data_from_slides(topic):
    """Extract vocabulary, examples, dialogues from slides for practice activities."""
    data = {
        "vocabulary": [],
        "examples": [],
        "dialogues": [],
        "questions": [],
        "mcq_questions": [],
    }

    if topic.get("slides_json"):
        try:
            obj = json.loads(topic["slides_json"])
            slides = obj.get("slides", obj) if isinstance(obj, dict) else obj
            for slide in slides:
                slide_type = slide.get("type", "")

                # Vocabulary
                if slide_type == "vocabulary" and slide.get("vocabulary"):
                    for v in slide["vocabulary"]:
                        if v.get("word") and v.get("meaning"):
                            data["vocabulary"].append({
                                "word": v["word"],
                                "meaning": v["meaning"],
                                "example": v.get("example", ""),
                            })

                # Examples
                if slide.get("examples"):
                    for ex in slide["examples"]:
                        if isinstance(ex, dict) and ex.get("en"):
                            data["examples"].append({"en": ex["en"], "th": ex.get("th", "")})
                        elif isinstance(ex, str):
                            data["examples"].append({"en": ex, "th": ""})

                # Dialogues
                if slide_type == "dialogue" and slide.get("lines"):
                    for line in slide["lines"]:
                        if isinstance(line, dict) and line.get("text"):
                            data["dialogues"].append({
                                "speaker": line.get("speaker", ""),
                                "text": line["text"],
                            })
        except Exception:
            pass

    # From game questions
    for set_no in range(1, 4):
        qs = GameQuestion.get_by_topic_and_set(topic["id"], set_no)
        for q in qs:
            data["questions"].append({"question": q["question"], "answer": q["answer"]})

    # From MCQ practice questions
    mcq_rows = _normalize_practice_questions(PracticeQuestion.get_by_topic(topic["id"]))
    data["mcq_questions"] = mcq_rows

    return data


# ---------------------------------------------------------------------------
# Sentence Builder helpers (AI Thai translation)
# ---------------------------------------------------------------------------
def _extract_first_json_array(s: str):
    if not s:
        return None
    s = s.strip()
    if s.startswith("[") and s.endswith("]"):
        try:
            return json.loads(s)
        except Exception:
            pass
    start = s.find("[")
    if start == -1:
        return None
    depth = 0
    for i in range(start, len(s)):
        ch = s[i]
        if ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                chunk = s[start : i + 1]
                try:
                    return json.loads(chunk)
                except Exception:
                    return None
    return None


def _ai_translate_en_to_th(sentences, model="gpt-4o-mini"):
    api_key = (
        os.environ.get("OPENAI_API_KEY")
        or os.environ.get("OPENAI_APIKEY")
        or os.environ.get("OPENAI_KEY")
    )
    if not api_key:
        return []
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
    except Exception:
        return []

    sys_msg = (
        "You translate English teaching examples into natural Thai. "
        "Return ONLY valid JSON array. No markdown. No extra text."
    )
    user_msg = {
        "task": "translate_en_to_th",
        "rules": [
            "Keep meaning faithful and natural for Thai students.",
            "Do not add explanations.",
            "Do not number items.",
            'Return format: [{"en":...,"th":...}, ...] in same order.',
        ],
        "sentences": sentences[:40],
    }
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": sys_msg},
                {"role": "user", "content": json.dumps(user_msg, ensure_ascii=False)},
            ],
            temperature=0.2,
        )
        content = (resp.choices[0].message.content or "").strip()
    except Exception:
        try:
            resp = client.responses.create(
                model=model,
                input=[
                    {"role": "system", "content": sys_msg},
                    {"role": "user", "content": json.dumps(user_msg, ensure_ascii=False)},
                ],
                temperature=0.2,
            )
            content = (resp.output_text or "").strip()
        except Exception:
            return []

    arr = _extract_first_json_array(content)
    if not isinstance(arr, list):
        return []
    out = []
    for it in arr:
        if isinstance(it, dict) and it.get("en") and it.get("th"):
            out.append({"en": str(it["en"]).strip(), "th": str(it["th"]).strip()})
    return out


def _sentence_builder_enrich_game_data_with_th(topic, game_data):
    """Ensure examples contain Thai prompts for Sentence Builder."""
    if not game_data:
        return game_data
    examples = game_data.get("examples") or []
    if not isinstance(examples, list) or not examples:
        return game_data

    need_en = []
    for ex in examples:
        if not isinstance(ex, dict):
            continue
        en = (ex.get("en") or "").strip()
        th = (ex.get("th") or "").strip()
        has_thai = bool(th and re.search(r"[\u0e01-\u0e59]", th))
        if en and not has_thai:
            if en not in need_en:
                need_en.append(en)

    if not need_en:
        return game_data

    translated = _ai_translate_en_to_th(need_en)
    if not translated:
        return game_data

    mapping = {
        str(t.get("en") or "").strip(): str(t.get("th") or "").strip()
        for t in translated
        if isinstance(t, dict)
    }

    new_examples = []
    for ex in examples:
        if isinstance(ex, dict) and ex.get("en"):
            en = str(ex.get("en") or "").strip()
            th = str(ex.get("th") or "").strip()
            has_thai = bool(th and re.search(r"[\u0e01-\u0e59]", th))
            if (not has_thai) and en in mapping and mapping[en]:
                th = mapping[en]
            new_examples.append({"en": en, "th": th})
        else:
            new_examples.append(ex)

    game_data["examples"] = new_examples
    return game_data


# ==============================================================================
# MCQ Practice Routes
# ==============================================================================
@practice_bp.route("/topic/<int:topic_id>/practice")
@login_required
def practice(topic_id):
    topic = _get_topic_or_404(topic_id)
    questions = _normalize_practice_questions(PracticeQuestion.get_by_topic(topic_id))
    link = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "mcq")
    student_url = (
        (request.url_root.rstrip("/") + url_for("practice.public_practice", token=link["token"]))
        if link
        else None
    )
    return render_template("practice.html", topic=topic, questions=questions, student_url=student_url)


@practice_bp.route("/api/practice/<int:topic_id>/submit", methods=["POST"])
@login_required
def api_practice_submit(topic_id):
    _get_topic_or_404(topic_id)
    data = request.get_json() or {}
    answers = data.get("answers", {})
    questions = _normalize_practice_questions(PracticeQuestion.get_by_topic(topic_id))
    score, total, feedback = 0, len(questions), {}
    for q in questions:
        qid = str(q["id"])
        ua = (answers.get(qid, "") or "").strip().lower()
        ca = (q.get("correct_answer") or "").strip().lower()
        correct = ua == ca
        if correct:
            score += 1
        feedback[qid] = {
            "is_correct": correct,
            "user_answer": answers.get(qid, ""),
            "correct_answer": q.get("correct_answer"),
        }
    pct = (score / total * 100) if total else 0
    AttemptHistory.create(session["user_id"], topic_id, score, total, pct)
    return jsonify({"score": score, "total": total, "percentage": pct, "feedback": feedback})


@practice_bp.route("/api/practice/<int:topic_id>/link", methods=["POST"])
@login_required
def api_practice_create_link(topic_id):
    _get_topic_or_404(topic_id)
    old = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "mcq")
    if not old:
        link = PracticeLink.create(topic_id, session["user_id"], secrets.token_urlsafe(12), "mcq")
    else:
        link = old
    return jsonify({"url": request.url_root.rstrip("/") + url_for("practice.public_practice", token=link["token"])})


@practice_bp.route("/topic/<int:topic_id>/practice/pdf")
@login_required
def practice_pdf(topic_id):
    topic = _get_topic_or_404(topic_id)
    include_answers = request.args.get("answers") == "1"
    pdf = _build_practice_pdf(
        topic["name"],
        _normalize_practice_questions(PracticeQuestion.get_by_topic(topic_id)),
        include_answers,
    )
    return (
        pdf,
        200,
        {
            "Content-Type": "application/pdf",
            "Content-Disposition": f"attachment; filename=practice_{topic_id}.pdf",
        },
    )


# ==============================================================================
# MCQ Scores
# ==============================================================================
@practice_bp.route("/topic/<int:topic_id>/practice/scores")
@login_required
def practice_scores(topic_id):
    topic = _get_topic_or_404(topic_id)
    from models import get_db

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "SELECT ps.* FROM practice_submissions ps "
        "JOIN practice_links pl ON ps.link_id=pl.id "
        "WHERE pl.topic_id=? AND pl.practice_type='mcq' ORDER BY ps.id DESC LIMIT 1000",
        (topic_id,),
    )
    submissions = [dict(r) for r in c.fetchall()]
    conn.close()
    classrooms = sorted(set(s.get("classroom") or "" for s in submissions if s.get("classroom")))
    return render_template("practice_scores.html", topic=topic, submissions=submissions, classrooms=classrooms)


@practice_bp.route("/topic/<int:topic_id>/practice/scores/csv")
@login_required
def practice_scores_csv(topic_id):
    topic = _get_topic_or_404(topic_id)
    from models import get_db

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "SELECT ps.* FROM practice_submissions ps "
        "JOIN practice_links pl ON ps.link_id=pl.id "
        "WHERE pl.topic_id=? AND pl.practice_type='mcq' ORDER BY ps.classroom,ps.student_no",
        (topic_id,),
    )
    rows = c.fetchall()
    conn.close()
    out = StringIO()
    w = csv.writer(out)
    w.writerow(["#", "Name", "No", "Class", "Score", "Total", "%", "Time"])
    for i, r in enumerate(rows, 1):
        w.writerow([
            i,
            r["student_name"],
            r["student_no"] or "",
            r["classroom"] or "",
            r["score"],
            r["total"],
            f"{r['percentage']:.0f}%",
            r["created_at"],
        ])
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=scores_{topic_id}.csv"},
    )


@practice_bp.route("/topic/<int:topic_id>/practice/scores/excel")
@login_required
def practice_scores_excel(topic_id):
    topic = _get_topic_or_404(topic_id)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
    except Exception:
        return redirect(url_for("practice.practice_scores_csv", topic_id=topic_id))

    from models import get_db

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "SELECT ps.* FROM practice_submissions ps "
        "JOIN practice_links pl ON ps.link_id=pl.id "
        "WHERE pl.topic_id=? AND pl.practice_type='mcq' ORDER BY ps.classroom,ps.student_no",
        (topic_id,),
    )
    rows = c.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="667eea")
    bd = Border(
        left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin")
    )
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Practice Scores: {topic['name']}"
    ws["A1"].font = Font(bold=True, size=14)
    for col, h in enumerate(["#", "Name", "No", "Class", "Score", "Total", "%", "Time"], 1):
        cell = ws.cell(3, col, h)
        cell.font, cell.fill, cell.border = hf, hfill, bd
    for i, r in enumerate(rows, 1):
        for col, v in enumerate(
            [
                i,
                r["student_name"],
                r["student_no"] or "",
                r["classroom"] or "",
                r["score"],
                r["total"],
                f"{r['percentage']:.0f}%",
                str(r["created_at"])[:19],
            ],
            1,
        ):
            cell = ws.cell(i + 3, col, v)
            cell.border = bd

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=scores_{topic_id}.xlsx"},
    )


# ==============================================================================
# Public MCQ Practice
# ==============================================================================
@practice_bp.route("/p/<token>")
def public_practice(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link.get("is_active"):
        return render_template("error.html", error_code=404, error_msg="\u0e25\u0e34\u0e07\u0e01\u0e4c\u0e2b\u0e21\u0e14\u0e2d\u0e32\u0e22\u0e38"), 404
    topic = Topic.get_by_id(link["topic_id"])
    if not topic:
        return render_template("error.html", error_code=404, error_msg="Topic not found"), 404
    classrooms = Classroom.get_by_owner(link["created_by"]) if link.get("created_by") else []
    return render_template(
        "practice_public.html",
        topic=topic,
        questions=_normalize_practice_questions(PracticeQuestion.get_by_topic(topic["id"])),
        token=token,
        classrooms=classrooms,
    )


@practice_bp.route("/api/p/<token>/submit", methods=["POST"])
def api_public_practice_submit(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link.get("is_active"):
        return jsonify({"error": "Invalid link"}), 404
    data = request.get_json() or {}
    name = (data.get("student_name") or "").strip()
    if not name:
        return jsonify({"error": "Name required"}), 400
    questions = _normalize_practice_questions(PracticeQuestion.get_by_topic(link["topic_id"]))
    answers = data.get("answers", {})
    score, total, feedback = 0, len(questions), {}
    for q in questions:
        qid = str(q["id"])
        ua = (answers.get(qid, "") or "").strip().lower()
        ca = (q.get("correct_answer") or "").strip().lower()
        correct = ua == ca
        if correct:
            score += 1
        feedback[qid] = {
            "is_correct": correct,
            "user_answer": answers.get(qid, ""),
            "correct_answer": q.get("correct_answer"),
        }
    pct = (score / total * 100) if total else 0
    PracticeSubmission.create(
        link["id"],
        name,
        data.get("student_no") or "",
        data.get("classroom") or "",
        json.dumps({"answers": answers}),
        score,
        total,
        pct,
    )
    return jsonify({"score": score, "total": total, "percentage": pct, "feedback": feedback})


# ==============================================================================
# Fill-in-the-Blanks
# ==============================================================================
@practice_bp.route("/topic/<int:topic_id>/practice/fill-blanks")
@login_required
def practice_fill_blanks(topic_id):
    topic = _get_topic_or_404(topic_id)
    practice_data = _get_practice_data_from_slides(topic)
    link = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "fill")
    student_url = None
    if link:
        student_url = request.url_root.rstrip("/") + url_for("practice.public_fill_blanks", token=link["token"])
    return render_template("practice_fill_blanks.html", topic=topic, practice_data=practice_data, student_url=student_url)


@practice_bp.route("/api/practice/<int:topic_id>/fill-blanks/link", methods=["POST"])
@login_required
def api_fill_blanks_create_link(topic_id):
    _get_topic_or_404(topic_id)
    old = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "fill")
    if not old:
        link = PracticeLink.create(topic_id, session["user_id"], secrets.token_urlsafe(12), "fill")
    else:
        link = old
    return jsonify({"url": request.url_root.rstrip("/") + url_for("practice.public_fill_blanks", token=link["token"])})


@practice_bp.route("/topic/<int:topic_id>/practice/fill-blanks/scores")
@login_required
def practice_fill_blanks_scores(topic_id):
    topic = _get_topic_or_404(topic_id)
    from models import get_db

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "SELECT ps.* FROM practice_submissions ps "
        "JOIN practice_links pl ON ps.link_id=pl.id "
        "WHERE pl.topic_id=? AND pl.practice_type='fill' ORDER BY ps.id DESC LIMIT 500",
        (topic_id,),
    )
    submissions = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template("practice_scores.html", topic=topic, submissions=submissions, practice_type="Fill in the Blanks")


@practice_bp.route("/p/fill/<token>")
def public_fill_blanks(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link["is_active"]:
        return "\u0e25\u0e34\u0e07\u0e01\u0e4c\u0e44\u0e21\u0e48\u0e16\u0e39\u0e01\u0e15\u0e49\u0e2d\u0e07\u0e2b\u0e23\u0e37\u0e2d\u0e2b\u0e21\u0e14\u0e2d\u0e32\u0e22\u0e38", 404
    topic = Topic.get_by_id(link["topic_id"])
    if not topic:
        return "Topic not found", 404
    practice_data = _get_practice_data_from_slides(topic)
    classrooms = Classroom.get_by_owner(link["created_by"]) if link.get("created_by") else []
    return render_template("practice_fill_blanks_public.html", topic=topic, practice_data=practice_data, token=token, classrooms=classrooms)


@practice_bp.route("/api/public/fill/<token>/submit", methods=["POST"])
def api_public_fill_blanks_submit(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link["is_active"]:
        return _json_error("Invalid link", 404)
    data = request.get_json() or {}
    student_name = (data.get("student_name") or data.get("name") or "Anonymous").strip()[:100]
    student_no = (data.get("student_no") or "").strip()[:20]
    classroom = (data.get("classroom") or "").strip()[:30]
    score = int(data.get("score", 0))
    total = int(data.get("total", 0))
    pct = (score / total * 100) if total else 0
    PracticeSubmission.create(
        link["id"], student_name, student_no, classroom,
        json.dumps(data.get("answers", {})), score, total, pct,
    )
    return jsonify({"ok": True, "score": score, "total": total, "percentage": pct})


# ==============================================================================
# Unscramble
# ==============================================================================
@practice_bp.route("/topic/<int:topic_id>/practice/unscramble")
@login_required
def practice_unscramble(topic_id):
    topic = _get_topic_or_404(topic_id)
    practice_data = _get_practice_data_from_slides(topic)
    link = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "unscramble")
    student_url = None
    if link:
        student_url = request.url_root.rstrip("/") + url_for("practice.public_unscramble", token=link["token"])
    return render_template("practice_unscramble.html", topic=topic, practice_data=practice_data, student_url=student_url)


@practice_bp.route("/api/practice/<int:topic_id>/unscramble/link", methods=["POST"])
@login_required
def api_unscramble_create_link(topic_id):
    _get_topic_or_404(topic_id)
    old = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "unscramble")
    if not old:
        link = PracticeLink.create(topic_id, session["user_id"], secrets.token_urlsafe(12), "unscramble")
    else:
        link = old
    return jsonify({"url": request.url_root.rstrip("/") + url_for("practice.public_unscramble", token=link["token"])})


@practice_bp.route("/topic/<int:topic_id>/practice/unscramble/scores")
@login_required
def practice_unscramble_scores(topic_id):
    topic = _get_topic_or_404(topic_id)
    from models import get_db

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "SELECT ps.* FROM practice_submissions ps "
        "JOIN practice_links pl ON ps.link_id=pl.id "
        "WHERE pl.topic_id=? AND pl.practice_type='unscramble' ORDER BY ps.id DESC LIMIT 500",
        (topic_id,),
    )
    submissions = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template("practice_scores.html", topic=topic, submissions=submissions, practice_type="Sentence Unscramble")


@practice_bp.route("/p/unscramble/<token>")
def public_unscramble(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link["is_active"]:
        return "\u0e25\u0e34\u0e07\u0e01\u0e4c\u0e44\u0e21\u0e48\u0e16\u0e39\u0e01\u0e15\u0e49\u0e2d\u0e07\u0e2b\u0e23\u0e37\u0e2d\u0e2b\u0e21\u0e14\u0e2d\u0e32\u0e22\u0e38", 404
    topic = Topic.get_by_id(link["topic_id"])
    if not topic:
        return "Topic not found", 404
    practice_data = _get_practice_data_from_slides(topic)
    classrooms = Classroom.get_by_owner(link["created_by"]) if link.get("created_by") else []
    return render_template("practice_unscramble_public.html", topic=topic, practice_data=practice_data, token=token, classrooms=classrooms)


@practice_bp.route("/api/public/unscramble/<token>/submit", methods=["POST"])
def api_public_unscramble_submit(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link["is_active"]:
        return _json_error("Invalid link", 404)
    data = request.get_json() or {}
    student_name = (data.get("student_name") or data.get("name") or "Anonymous").strip()[:100]
    student_no = (data.get("student_no") or "").strip()[:20]
    classroom = (data.get("classroom") or "").strip()[:30]
    score = int(data.get("score", 0))
    total = int(data.get("total", 0))
    pct = (score / total * 100) if total else 0
    PracticeSubmission.create(
        link["id"], student_name, student_no, classroom,
        json.dumps(data.get("answers", {})), score, total, pct,
    )
    return jsonify({"ok": True, "score": score, "total": total, "percentage": pct})


# ==============================================================================
# QR Short URLs (redirect to public practice pages)
# ==============================================================================
@practice_bp.route("/p/<int:topic_id>/mcq")
def qr_practice_mcq(topic_id):
    topic = Topic.get_by_id(topic_id)
    if not topic:
        abort(404)
    link = PracticeLink.get_by_topic(topic_id)
    if not link:
        link = PracticeLink.create(topic_id, topic.get("owner_id") or 1, secrets.token_urlsafe(12), "mcq")
    return redirect(url_for("practice.public_practice", token=link["token"]))


@practice_bp.route("/p/<int:topic_id>/fill")
def qr_practice_fill(topic_id):
    topic = Topic.get_by_id(topic_id)
    if not topic:
        abort(404)
    link = PracticeLink.get_by_topic(topic_id)
    if not link:
        link = PracticeLink.create(topic_id, topic.get("owner_id") or 1, secrets.token_urlsafe(12), "fill")
    return redirect(url_for("practice.public_fill_blanks", token=link["token"]))


@practice_bp.route("/p/<int:topic_id>/unscramble")
def qr_practice_unscramble(topic_id):
    topic = Topic.get_by_id(topic_id)
    if not topic:
        abort(404)
    link = PracticeLink.get_by_topic(topic_id)
    if not link:
        link = PracticeLink.create(topic_id, topic.get("owner_id") or 1, secrets.token_urlsafe(12), "unscramble")
    return redirect(url_for("practice.public_unscramble", token=link["token"]))


# ==============================================================================
# Self-Study
# ==============================================================================
@practice_bp.route("/api/topic/<int:topic_id>/study/link", methods=["POST"])
@login_required
def api_create_study_link(topic_id):
    topic = _get_topic_or_404(topic_id)
    token = secrets.token_urlsafe(16)
    link = PracticeLink.create(topic_id, session["user_id"], token, "self_study")
    study_url = request.url_root.rstrip("/") + url_for("practice.public_self_study", token=token)
    return jsonify({"success": True, "url": study_url, "token": token})


@practice_bp.route("/api/study/<token>/submit", methods=["POST"])
def api_study_submit(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link.get("is_active"):
        return jsonify({"error": "Invalid link"}), 404
    data = request.get_json() or {}
    name = (data.get("student_name") or "").strip()
    if not name:
        return jsonify({"error": "Name required"}), 400
    score = int(data.get("score", 0))
    total = int(data.get("total", 0))
    pct = float(data.get("percentage", 0))
    practice_type = data.get("practice_type", "self_study")
    PracticeSubmission.create(
        link_id=link["id"],
        student_name=name,
        student_no=data.get("student_no") or "",
        classroom=data.get("classroom") or "",
        answers_json=json.dumps({"type": practice_type, "submitted_at": datetime.now().isoformat()}),
        score=score,
        total=total,
        percentage=pct,
    )
    return jsonify({"success": True, "score": score, "total": total, "percentage": pct})


@practice_bp.route("/study/<token>")
def public_self_study(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link.get("is_active"):
        return render_template("error.html", error_code=404, error_msg="\u0e25\u0e34\u0e07\u0e01\u0e4c\u0e44\u0e21\u0e48\u0e16\u0e39\u0e01\u0e15\u0e49\u0e2d\u0e07\u0e2b\u0e23\u0e37\u0e2d\u0e2b\u0e21\u0e14\u0e2d\u0e32\u0e22\u0e38"), 404
    topic = Topic.get_by_id(link["topic_id"])
    if not topic:
        return render_template("error.html", error_code=404, error_msg="\u0e44\u0e21\u0e48\u0e1e\u0e1a\u0e1a\u0e17\u0e40\u0e23\u0e35\u0e22\u0e19"), 404

    slides = []
    if topic.get("slides_json"):
        try:
            obj = json.loads(topic["slides_json"])
            slides = obj.get("slides", obj) if isinstance(obj, dict) else obj
        except Exception:
            slides = []

    practice_data = {"mcq": [], "fill": []}
    mcq_questions = _normalize_practice_questions(PracticeQuestion.get_by_topic(topic["id"]))
    practice_data["mcq"] = mcq_questions
    practice_data["fill"] = _get_practice_data_from_slides(topic)

    classrooms = Classroom.get_by_owner(link["created_by"]) if link.get("created_by") else []

    return render_template(
        "self_study.html",
        topic=topic,
        token=token,
        slides_json=json.dumps(slides, ensure_ascii=False),
        practice_json=json.dumps(practice_data, ensure_ascii=False),
        classrooms=classrooms,
    )
