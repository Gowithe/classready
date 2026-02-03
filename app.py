# ==============================================================================
# FILE: app.py
# Teacher Platform MVP (Flask + SQLite)
# UPDATED: Classroom Management + Assignments + Game Sessions + Practice Export
# FIXED: Sentence Builder Syntax Error
# ==============================================================================

import os
import json
import secrets
import traceback
import base64
import csv
import re
from io import BytesIO, StringIO
from functools import wraps
from datetime import datetime
from typing import Optional, Dict, Any, List

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, send_from_directory, abort, Response
)
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename

from models import (
    get_db, init_db, User, Topic, GameQuestion, PracticeQuestion, AttemptHistory,
    PracticeLink, PracticeSubmission, GameSession, Classroom, ClassroomStudent, Assignment,
    LibrarySubject, LibraryUnit, UserSubscription, LibraryClone, LibraryRating, SubscriptionPlan,
    UsageLimits, PaymentTransaction,
)
from ai_generator import generate_lesson_bundle

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.utils import simpleSplit

from functools import wraps
from flask import abort, session
# from models import (..., PaymentTransaction)  # <-- INVALID (old pasted line). Kept as comment.
import requests
import urllib.parse
from dotenv import load_dotenv
load_dotenv()  # โหลดค่าจากไฟล์ .env เข้า os.environ

init_db()  # ensure DB tables exist (Render + Persistent Disk)
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key-change-in-prod")


# ==============================================================================
# Email (Gmail SMTP) - Email verification
# Set these env vars:
#   GMAIL_USER="your@gmail.com"
#   GMAIL_APP_PASSWORD="16-char app password"
#   APP_BASE_URL="http://127.0.0.1:5000"  (optional; for production set https://yourdomain)
# ==============================================================================
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

GMAIL_USER = os.environ.get("GMAIL_USER", "").strip()
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "").strip()
APP_BASE_URL = os.environ.get("APP_BASE_URL", "").strip()

def _build_external_url(path: str) -> str:
    if APP_BASE_URL:
        return APP_BASE_URL.rstrip("/") + path
    # fallback: Flask will build with current request context (may be http://127.0.0.1)
    return path

def send_verify_email(to_email: str, verify_link: str) -> None:
    """Send email verification link via Gmail SMTP (STARTTLS)."""
    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        # Dev fallback: no SMTP configured
        print("[EMAIL] SMTP not configured. VERIFY LINK:", verify_link)
        return

    msg = MIMEMultipart("alternative")
    msg["From"] = GMAIL_USER
    msg["To"] = to_email
    msg["Subject"] = "ยืนยันอีเมลเพื่อเข้าใช้งาน Teacher Platform"

    html = f"""    <div style="font-family:Arial,sans-serif;line-height:1.6">
      <h2>ยืนยันอีเมล</h2>
      <p>ขอบคุณที่สมัครใช้งาน กรุณาคลิกลิงก์ด้านล่างเพื่อยืนยันอีเมลของคุณ</p>
      <p><a href="{verify_link}" style="display:inline-block;padding:10px 14px;background:#667eea;color:#fff;text-decoration:none;border-radius:10px">ยืนยันอีเมล</a></p>
      <p style="color:#64748b;font-size:13px">ถ้าปุ่มกดไม่ได้ ให้คัดลอกลิงก์นี้ไปวางในเบราว์เซอร์:</p>
      <p style="word-break:break-all">{verify_link}</p>
      <p style="color:#64748b;font-size:13px">ลิงก์นี้มีอายุ 24 ชั่วโมง</p>
    </div>
    """

    msg.attach(MIMEText(html, "html", "utf-8"))

    with smtplib.SMTP("smtp.gmail.com", 587, timeout=20) as server:
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.send_message(msg)



# Payment Config - แก้ไขตามข้อมูลของคุณ
PROMPTPAY_ID = os.environ.get("PROMPTPAY_ID", "1234567890123")  # เลขบัตรประชาชน 13 หลัก
PROMPTPAY_NAME = os.environ.get("PROMPTPAY_NAME", "ชื่อบัญชี")
EASYSLIP_API_KEY = os.environ.get("EASYSLIP_API_KEY", "your_api_key")

# -----------------------------------------------------------------------------
# SQLite on Render Persistent Disk (recommended for now)
# - Add a Disk on Render (e.g., mount path: /var/data)
# - Set env var SQLITE_PATH=/var/data/teacher_platform.db
# The database file will then persist across deploys/restarts.
# -----------------------------------------------------------------------------

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024
ALLOWED_EXTENSIONS = {"pdf"}
ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}

def allowed_file(filename): return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS
def allowed_image(filename): return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS

with app.app_context():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@teacherplatform.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Admin@12345")
    if not User.get_by_email(admin_email):
        User.create(admin_email, admin_password, "admin")

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in first.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not _is_admin():
            abort(403)
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in first.", "error")
            return redirect(url_for("login"))
        user = User.get_by_id(session["user_id"])
        if not user or user.get("role") != "admin":
            flash("Admin access required.", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated

def _is_admin(): return session.get("role") == "admin"
def _can_access_topic(topic): return _is_admin() or int(topic.get("owner_id") or 0) == int(session.get("user_id") or 0)
def _get_topic_or_404(topic_id):
    topic = Topic.get_by_id(topic_id)
    if not topic: abort(404)
    if not _can_access_topic(topic): abort(403)
    return topic

def _wants_json_response(): return request.path.startswith("/api/") or "application/json" in (request.headers.get("Accept") or "").lower()
def _json_error(message, status=400): return jsonify({"ok": False, "error": message}), status

# ==============================================================================
# Freemium Helpers (added by app_patches.py) - keeps old code intact
# ==============================================================================
def is_premium_user(user_id: int) -> bool:
    # ตรวจสอบว่า user เป็น Premium หรือไม่
    try:
        return UserSubscription.is_premium(user_id)
    except Exception:
        return False

@app.context_processor
def inject_freemium_data():
    # Inject freemium data to all templates
    if "user_id" in session:
        try:
            user_id = session["user_id"]
            is_premium = is_premium_user(user_id)
            stats = UsageLimits.get_user_stats(user_id)
            return {
                "user_is_premium": is_premium,
                "usage_stats": stats,
                "free_limits": {
                    "topics": UsageLimits.FREE_TOPICS,
                    "classrooms": UsageLimits.FREE_CLASSROOMS,
                    "ai_generate": UsageLimits.FREE_AI_GENERATE_PER_MONTH,
                    "students_per_classroom": UsageLimits.FREE_STUDENTS_PER_CLASSROOM,
                },
            }
        except Exception as e:
            print(f"Context processor error: {e}")
    return {
        "user_is_premium": False,
        "usage_stats": {"topic_count": 0, "classroom_count": 0, "ai_generate_count": 0},
        "free_limits": {
            "topics": 5,
            "classrooms": 2,
            "ai_generate": 3,
            "students_per_classroom": 50,
        },
    }



# ==============================================================================
# Auth & Landing
# ==============================================================================
@app.route("/")
def landing(): return render_template("landing.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = (request.form.get("password") or "").strip()
        confirm = (request.form.get("confirm_password") or "").strip()

        if not email or not password:
            flash("Email and password required.", "error")
            return render_template("register.html")
        if password != confirm:
            flash("Passwords do not match.", "error")
            return render_template("register.html")
        if User.get_by_email(email):
            flash("Email already registered.", "error")
            return render_template("register.html")

        # Create user as unverified + generate verify_token
        user = User.create(email, password, "teacher")
        token = (user or {}).get("verify_token")
        if not token:
            flash("Could not create verification token. Please try again.", "error")
            return render_template("register.html")

        verify_path = url_for("verify_email", token=token)
        verify_link = _build_external_url(verify_path) if APP_BASE_URL else url_for("verify_email", token=token, _external=True)

        try:
            send_verify_email(email, verify_link)
        except Exception:
            traceback.print_exc()
            flash("ส่งอีเมลไม่สำเร็จ กรุณาลองใหม่ภายหลัง", "error")
            return render_template("register.html")

        return render_template("verify_sent.html", email=email)

    return render_template("register.html")




@app.route("/verify/<token>")
def verify_email(token):
    # Email verification endpoint
    user = User.get_by_verify_token(token)
    if not user:
        flash("ลิงก์ไม่ถูกต้องหรือถูกใช้งานไปแล้ว", "error")
        return redirect(url_for("login"))

    # Expiry check (UTC)
    try:
        exp = user.get("verify_expires")
        if exp:
            exp_dt = datetime.fromisoformat(exp)
            if datetime.utcnow() > exp_dt:
                flash("ลิงก์ยืนยันหมดอายุ กรุณาสมัครใหม่หรือขอส่งลิงก์อีกครั้ง", "error")
                return redirect(url_for("login"))
    except Exception:
        pass

    User.mark_verified(user["id"])
    flash("ยืนยันอีเมลสำเร็จ! กรุณาเข้าสู่ระบบ", "success")
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = (request.form.get("password") or "").strip()
        user = User.get_by_email(email)
        if user and check_password_hash(user["password_hash"], password):
            # Require email verification
            if not int(user.get("is_verified") or 0):
                flash("กรุณายืนยันอีเมลก่อนเข้าใช้งาน (ตรวจสอบในกล่องจดหมาย)", "error")
                return redirect(url_for("login"))
            session["user_id"] = user["id"]
            session["email"] = user["email"]
            session["role"] = user["role"]
            return redirect(url_for("dashboard"))
        flash("Invalid email or password.", "error")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("landing"))


# ==============================================================================
# Resend Verification Email
# ==============================================================================
@app.route("/resend-verification", methods=["GET", "POST"])
def resend_verification():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        
        if not email:
            flash("กรุณากรอกอีเมล", "error")
            return render_template("resend_verification.html")
        
        user = User.get_by_email(email)
        
        if not user:
            flash("ไม่พบอีเมลนี้ในระบบ", "error")
            return render_template("resend_verification.html")
        
        if user.get("is_verified"):
            flash("อีเมลนี้ได้รับการยืนยันแล้ว สามารถเข้าสู่ระบบได้เลย", "success")
            return redirect(url_for("login"))
        
        # Refresh token and send email
        token = User.refresh_verify_token(user["id"])
        if token:
            verify_path = url_for("verify_email", token=token)
            verify_link = _build_external_url(verify_path) if APP_BASE_URL else url_for("verify_email", token=token, _external=True)
            try:
                send_verify_email(email, verify_link)
                flash("ส่งลิงก์ยืนยันใหม่ไปที่อีเมลของคุณแล้ว กรุณาตรวจสอบกล่องจดหมาย", "success")
            except Exception as e:
                print(f"[EMAIL ERROR] {e}")
                flash("ไม่สามารถส่งอีเมลได้ กรุณาลองใหม่อีกครั้ง", "error")
        
        return redirect(url_for("login"))
    
    return render_template("resend_verification.html")


# ==============================================================================
# Forgot Password
# ==============================================================================
@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        
        if not email:
            flash("กรุณากรอกอีเมล", "error")
            return render_template("forgot_password.html")
        
        user = User.get_by_email(email)
        
        # Always show success message (security: don't reveal if email exists)
        if user:
            token = User.set_reset_token(user["id"])
            if token:
                reset_path = url_for("reset_password", token=token)
                reset_link = _build_external_url(reset_path) if APP_BASE_URL else url_for("reset_password", token=token, _external=True)
                try:
                    send_reset_password_email(email, reset_link)
                except Exception as e:
                    print(f"[EMAIL ERROR] {e}")
        
        flash("หากอีเมลนี้มีอยู่ในระบบ เราได้ส่งลิงก์รีเซ็ตรหัสผ่านไปแล้ว กรุณาตรวจสอบกล่องจดหมาย", "success")
        return redirect(url_for("login"))
    
    return render_template("forgot_password.html")


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    user = User.get_by_reset_token(token)
    
    if not user:
        flash("ลิงก์รีเซ็ตรหัสผ่านไม่ถูกต้องหรือหมดอายุแล้ว", "error")
        return redirect(url_for("forgot_password"))
    
    if request.method == "POST":
        password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")
        
        if len(password) < 6:
            flash("รหัสผ่านต้องมีอย่างน้อย 6 ตัวอักษร", "error")
            return render_template("reset_password.html", token=token)
        
        if password != confirm:
            flash("รหัสผ่านไม่ตรงกัน", "error")
            return render_template("reset_password.html", token=token)
        
        User.reset_password(user["id"], password)
        flash("รีเซ็ตรหัสผ่านสำเร็จ! สามารถเข้าสู่ระบบด้วยรหัสผ่านใหม่ได้แล้ว", "success")
        return redirect(url_for("login"))
    
    return render_template("reset_password.html", token=token)


# ==============================================================================
# My Account
# ==============================================================================
@app.route("/my-account")
@login_required
def my_account():
    user = User.get_by_id(session["user_id"])
    if not user:
        return redirect(url_for("logout"))
    
    # Get subscription info
    subscription = UserSubscription.get_active_subscription(user["id"])
    is_premium = subscription is not None
    
    # Get usage stats from UsageLimits
    raw_stats = UsageLimits.get_user_stats(user["id"])
    
    # Count students
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT COUNT(*) FROM classroom_students cs
        JOIN classrooms cl ON cs.classroom_id = cl.id
        WHERE cl.owner_id = ?
    """, (user["id"],))
    student_count = c.fetchone()[0]
    conn.close()
    
    # Map to template-expected keys
    stats = {
        "topics": raw_stats.get("topic_count", 0),
        "classrooms": raw_stats.get("classroom_count", 0),
        "ai_this_month": raw_stats.get("ai_generate_count", 0),
        "students": student_count,
    }
    
    # Get limits
    if is_premium:
        limits = {
            "topics": None,  # None = unlimited
            "classrooms": None,
            "ai_generate": None,
            "students_per_classroom": None,
        }
    else:
        limits = {
            "topics": UsageLimits.FREE_TOPICS,
            "classrooms": UsageLimits.FREE_CLASSROOMS,
            "ai_generate": UsageLimits.FREE_AI_GENERATE_PER_MONTH,
            "students_per_classroom": UsageLimits.FREE_STUDENTS_PER_CLASSROOM,
        }
    
    return render_template("my_account.html", 
                         user=user, 
                         subscription=subscription, 
                         is_premium=is_premium,
                         stats=stats,
                         limits=limits)


@app.route("/my-account/update", methods=["POST"])
@login_required
def my_account_update():
    display_name = (request.form.get("display_name") or "").strip()[:100]
    User.update_profile(session["user_id"], display_name)
    flash("อัปเดตข้อมูลเรียบร้อย", "success")
    return redirect(url_for("my_account"))


@app.route("/my-account/change-password", methods=["POST"])
@login_required
def my_account_change_password():
    user = User.get_by_id(session["user_id"])
    if not user:
        return redirect(url_for("logout"))
    
    current_password = request.form.get("current_password", "")
    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")
    
    # Verify current password
    if not check_password_hash(user["password_hash"], current_password):
        flash("รหัสผ่านปัจจุบันไม่ถูกต้อง", "error")
        return redirect(url_for("my_account"))
    
    if len(new_password) < 6:
        flash("รหัสผ่านใหม่ต้องมีอย่างน้อย 6 ตัวอักษร", "error")
        return redirect(url_for("my_account"))
    
    if new_password != confirm_password:
        flash("รหัสผ่านใหม่ไม่ตรงกัน", "error")
        return redirect(url_for("my_account"))
    
    User.update_password(session["user_id"], new_password)
    flash("เปลี่ยนรหัสผ่านสำเร็จ!", "success")
    return redirect(url_for("my_account"))


# ==============================================================================
# Send Reset Password Email
# ==============================================================================
def send_reset_password_email(to_email: str, reset_link: str) -> None:
    """Send password reset link via Gmail SMTP."""
    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        print("[EMAIL] SMTP not configured. RESET LINK:", reset_link)
        return

    msg = MIMEMultipart("alternative")
    msg["From"] = GMAIL_USER
    msg["To"] = to_email
    msg["Subject"] = "รีเซ็ตรหัสผ่าน - Teacher Platform"

    html = f"""
    <div style="font-family:Arial,sans-serif;line-height:1.6;max-width:500px;margin:0 auto;">
      <h2 style="color:#667eea;">🔐 รีเซ็ตรหัสผ่าน</h2>
      <p>คุณได้ร้องขอรีเซ็ตรหัสผ่านสำหรับบัญชี Teacher Platform</p>
      <p>คลิกปุ่มด้านล่างเพื่อตั้งรหัสผ่านใหม่:</p>
      <p style="text-align:center;margin:24px 0;">
        <a href="{reset_link}" style="display:inline-block;padding:12px 24px;background:linear-gradient(135deg,#667eea,#764ba2);color:#fff;text-decoration:none;border-radius:10px;font-weight:bold;">
          ตั้งรหัสผ่านใหม่
        </a>
      </p>
      <p style="color:#64748b;font-size:13px;">ถ้าปุ่มกดไม่ได้ ให้คัดลอกลิงก์นี้ไปวางในเบราว์เซอร์:</p>
      <p style="word-break:break-all;font-size:13px;background:#f1f5f9;padding:10px;border-radius:6px;">{reset_link}</p>
      <p style="color:#ef4444;font-size:13px;">⚠️ ลิงก์นี้มีอายุ 1 ชั่วโมง</p>
      <hr style="border:none;border-top:1px solid #e2e8f0;margin:20px 0;">
      <p style="color:#94a3b8;font-size:12px;">หากคุณไม่ได้ร้องขอรีเซ็ตรหัสผ่าน กรุณาเพิกเฉยอีเมลนี้</p>
    </div>
    """

    msg.attach(MIMEText(html, "html", "utf-8"))

    with smtplib.SMTP("smtp.gmail.com", 587, timeout=20) as server:
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.send_message(msg)


@app.route("/uploads/<path:filename>")
@login_required
def uploaded_file(filename): return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


# ==============================================================================
# Dashboard
# ==============================================================================
@app.route("/dashboard")
@login_required
def dashboard():
    user_id = session["user_id"]
    my_topics = Topic.get_by_owner(user_id)
    all_topics = Topic.get_all() if _is_admin() else my_topics
    recent = AttemptHistory.get_recent_by_user(user_id, limit=5)
    classrooms = Classroom.get_by_owner(user_id)
    
    # ========== Dashboard Statistics ==========
    
    # Basic stats
    stats = {
        "total_topics": len(my_topics),
        "total_classrooms": len(classrooms),
        "total_students": 0,
        "total_submissions": 0
    }
    
    # Count total students
    for c in classrooms:
        students = ClassroomStudent.get_by_classroom(c["id"])
        stats["total_students"] += len(students)
    
    # Classroom progress & submissions
    classroom_progress = []
    all_submissions = []
    
    for c in classrooms:
        students = ClassroomStudent.get_by_classroom(c["id"])
        assignments = Assignment.get_by_classroom(c["id"])
        
        total_students = len(students)
        submitted_count = 0
        
        for a in assignments:
            status = Assignment.get_submissions_status(a["id"])
            submitted_count += len(status.get("submitted", []))
            all_submissions.extend(status.get("submissions", []))
        
        # Calculate submission rate
        total_possible = total_students * len(assignments) if assignments else 0
        percentage = int((submitted_count / total_possible * 100)) if total_possible > 0 else 0
        
        classroom_progress.append({
            "id": c["id"],
            "name": c["name"],
            "total": total_students,
            "submitted": submitted_count,
            "percentage": percentage
        })
    
    stats["total_submissions"] = len(all_submissions)
    
    # ========== Alerts ==========
    alerts = []
    
    for c in classrooms:
        assignments = Assignment.get_by_classroom(c["id"])
        for a in assignments:
            status = Assignment.get_submissions_status(a["id"])
            not_submitted = status.get("not_submitted", [])
            
            # Alert: Students who haven't submitted
            if not_submitted:
                alerts.append({
                    "type": "warning",
                    "title": f"{len(not_submitted)} คนยังไม่ส่งงาน",
                    "message": f"งาน '{a['title']}' ห้อง {c['name']}"
                })
            
            # Alert: Low scores
            for sub in status.get("submissions", []):
                if sub.get("percentage", 100) < 50:
                    alerts.append({
                        "type": "danger",
                        "title": f"{sub.get('student_name', 'นักเรียน')} คะแนนต่ำ",
                        "message": f"ได้ {sub.get('percentage', 0):.0f}% ในงาน '{a['title']}'"
                    })
    
    # ========== Top & Struggling Students ==========
    student_scores = {}  # {name: {classroom, scores: [], avg}}
    
    for sub in all_submissions:
        name = sub.get("student_name", "").strip()
        if not name:
            continue
        
        if name not in student_scores:
            student_scores[name] = {
                "name": name,
                "classroom": sub.get("classroom", "-"),
                "scores": []
            }
        student_scores[name]["scores"].append(sub.get("percentage", 0))
    
    # Calculate averages
    for name, data in student_scores.items():
        if data["scores"]:
            data["avg_score"] = int(sum(data["scores"]) / len(data["scores"]))
        else:
            data["avg_score"] = 0
    
    # Sort for top performers and struggling students
    sorted_students = sorted(student_scores.values(), key=lambda x: x["avg_score"], reverse=True)
    top_students = [s for s in sorted_students if s["avg_score"] >= 70][:5]
    struggling_students = [s for s in sorted_students if s["avg_score"] < 50][:5]
    
    return render_template(
        "dashboard.html",
        my_topics=my_topics,
        topics=all_topics,
        recent=recent,
        classrooms=classrooms,
        stats=stats,
        classroom_progress=classroom_progress,
        alerts=alerts[:20],  # Limit alerts
        top_students=top_students,
        struggling_students=struggling_students
    )

@app.route("/topic/<int:topic_id>")
@login_required
def topic_detail(topic_id):
    topic = _get_topic_or_404(topic_id)
    AttemptHistory.track_view(session["user_id"], topic_id)
    is_owner = int(topic.get("owner_id") or 0) == int(session["user_id"])
    has_game = len(GameQuestion.get_by_topic_and_set(topic_id, 1) or []) > 0
    has_practice = len(PracticeQuestion.get_by_topic(topic_id) or []) > 0
    has_slides = False
    if topic.get("slides_json"):
        try:
            obj = json.loads(topic["slides_json"])
            slides = obj.get("slides", obj) if isinstance(obj, dict) else obj
            has_slides = len(slides) > 0
        except: pass
    return render_template("topic_detail.html", topic=topic, is_owner=is_owner, is_admin=_is_admin(), has_game=has_game, has_practice=has_practice, has_slides=has_slides)


# ==============================================================================
# My Topics CRUD
# ==============================================================================
@app.route("/my/topics/create", methods=["GET", "POST"])
@login_required
def my_create_topic():
    # ตรวจสอบ limit (Freemium)
    is_premium = is_premium_user(session["user_id"])
    can_create, msg = UsageLimits.can_create_topic(session["user_id"], is_premium)

    if request.method == "POST":
        if not can_create:
            flash(f"❌ {msg} - อัปเกรดเป็น Premium เพื่อสร้างไม่จำกัด!", "error")
            return redirect(url_for("pricing"))

        name = (request.form.get("name") or "").strip()
        description = (request.form.get("description") or "").strip()
        if not name:
            flash("Topic name required.", "error")
            return render_template("my_topic_edit.html", topic=None, mode="create", can_create=can_create, limit_msg=msg)

        topic = Topic.create(session["user_id"], name, description, json.dumps({"slides": []}, ensure_ascii=False), "manual", None)
        flash("✅ สร้าง Topic สำเร็จ!", "success")
        return redirect(url_for("my_edit_topic", topic_id=topic["id"]))

    return render_template("my_topic_edit.html", topic=None, mode="create", can_create=can_create, limit_msg=msg)

@app.route("/my/topics/<int:topic_id>/edit", methods=["GET", "POST"])
@login_required
def my_edit_topic(topic_id):
    topic = _get_topic_or_404(topic_id)
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        description = (request.form.get("description") or "").strip()
        slides_json = (request.form.get("slides_json") or "").strip()
        if not name:
            flash("Topic name required.", "error")
            return render_template("my_topic_edit.html", topic=topic, mode="edit")
        try: json.loads(slides_json)
        except:
            flash("Invalid JSON.", "error")
            return render_template("my_topic_edit.html", topic=topic, mode="edit")
        pdf_filename = topic.get("pdf_file")
        file = request.files.get("pdf_file")
        if file and file.filename and allowed_file(file.filename):
            safe_name = secure_filename(file.filename)
            final_name = f"user{session['user_id']}_topic{topic_id}_{secrets.token_hex(6)}_{safe_name}"
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], final_name))
            pdf_filename = final_name
        Topic.update(topic_id, name, description, slides_json, pdf_filename)
        flash("Saved.", "success")
        return redirect(url_for("topic_detail", topic_id=topic_id))
    return render_template("my_topic_edit.html", topic=topic, mode="edit")

@app.route("/my/topics/<int:topic_id>/delete", methods=["POST"])
@login_required
def my_delete_topic(topic_id):
    _get_topic_or_404(topic_id)
    Topic.delete(topic_id)
    return redirect(url_for("dashboard"))


# ==============================================================================
# Slides
# ==============================================================================
@app.route("/topic/<int:topic_id>/slides")
@login_required
def view_slides(topic_id):
    topic = _get_topic_or_404(topic_id)
    
    # Check if topic has generated slides first
    slides = []
    if topic.get("slides_json"):
        try:
            obj = json.loads(topic["slides_json"])
            slides = obj.get("slides", obj) if isinstance(obj, dict) else obj
        except: pass
    
    # If has slides, show slides viewer
    if slides:
        return render_template("slides_viewer.html", topic=topic, slides=slides)
    
    # If no slides but has PDF, show PDF presentation
    if topic.get("pdf_file"):
        return render_template("slides_pdf_presentation.html", topic=topic, pdf_url=url_for("uploaded_file", filename=topic["pdf_file"]))
    
    # No slides and no PDF - show empty slides viewer
    return render_template("slides_viewer.html", topic=topic, slides=[])

@app.route("/topic/<int:topic_id>/slides/edit")
@login_required
def edit_slides(topic_id):
    topic = _get_topic_or_404(topic_id)
    return render_template("slides_editor.html", topic=topic)

@app.route("/api/topic/<int:topic_id>/slides", methods=["POST"])
@login_required
def api_save_slides(topic_id):
    topic = _get_topic_or_404(topic_id)
    data = request.get_json(silent=True) or {}
    slides = data.get("slides", [])
    processed = []
    for i, slide in enumerate(slides):
        ps = dict(slide)
        img_url = slide.get("image_url", "")
        if img_url and img_url.startswith("data:image"):
            try:
                header, b64 = img_url.split(",", 1)
                ext = "png" if "png" in header else "gif" if "gif" in header else "jpg"
                fn = f"slide_img_{topic_id}_{i}_{secrets.token_hex(6)}.{ext}"
                with open(os.path.join(app.config["UPLOAD_FOLDER"], fn), "wb") as f:
                    f.write(base64.b64decode(b64))
                ps["image_url"] = url_for("uploaded_file", filename=fn)
            except: pass
        processed.append(ps)
    Topic.update(topic_id, topic["name"], topic["description"], json.dumps({"slides": processed}, ensure_ascii=False), topic.get("pdf_file"))
    return jsonify({"ok": True})


# ==============================================================================
# Download Slides as PDF
# ==============================================================================
@app.route("/topic/<int:topic_id>/slides/download")
@login_required
def download_slides_pdf(topic_id):
    topic = _get_topic_or_404(topic_id)
    
    # Parse slides
    slides = []
    if topic.get("slides_json"):
        try:
            obj = json.loads(topic["slides_json"])
            slides = obj.get("slides", obj) if isinstance(obj, dict) else obj
        except: pass
    
    if not slides:
        flash("ไม่มีสไลด์", "error")
        return redirect(url_for("topic_detail", topic_id=topic_id))
    
    # Generate PDF
    pdf_bytes = _generate_slides_pdf(topic["name"], slides)
    
    # Clean filename
    safe_name = "".join(c for c in topic["name"] if c.isalnum() or c in " -_").strip()[:50] or "slides"
    
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={safe_name}_slides.pdf"}
    )


def _generate_slides_pdf(title, slides):
    # Generate a PDF from slides data - supports all slide types + Thai language
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import textwrap
    import urllib.request
    import tempfile
    
    buf = BytesIO()
    page_size = landscape(A4)
    c = canvas.Canvas(buf, pagesize=page_size)
    w, h = page_size
    
    # Register Thai font
    thai_font = "Helvetica"
    thai_font_bold = "Helvetica-Bold"
    thai_font_italic = "Helvetica-Oblique"
    
    # Try to find and register a Thai-compatible font
    font_paths = [
        # Windows fonts
        "C:/Windows/Fonts/THSarabunNew.ttf",
        "C:/Windows/Fonts/thsarabunnew.ttf",
        "C:/Windows/Fonts/Tahoma.ttf",
        "C:/Windows/Fonts/tahoma.ttf",
        "C:/Windows/Fonts/cordia.ttf",
        "C:/Windows/Fonts/CordiaNew.ttf",
        "C:/Windows/Fonts/angsana.ttf",
        # Linux fonts
        "/usr/share/fonts/truetype/thai/TH Sarabun New.ttf",
        "/usr/share/fonts/truetype/tlwg/TlwgTypo.ttf",
        "/usr/share/fonts/truetype/noto/NotoSansThai-Regular.ttf",
        # Mac fonts
        "/Library/Fonts/Thonburi.ttf",
        "/System/Library/Fonts/Thonburi.ttc",
    ]
    
    font_bold_paths = [
        "C:/Windows/Fonts/THSarabunNew Bold.ttf",
        "C:/Windows/Fonts/thsarabunnew-bold.ttf",
        "C:/Windows/Fonts/tahomabd.ttf",
        "C:/Windows/Fonts/cordiab.ttf",
    ]
    
    for fp in font_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("ThaiFont", fp))
                thai_font = "ThaiFont"
                thai_font_italic = "ThaiFont"
                print(f"Registered Thai font: {fp}")
                break
            except Exception as e:
                print(f"Failed to register font {fp}: {e}")
    
    for fp in font_bold_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("ThaiFontBold", fp))
                thai_font_bold = "ThaiFontBold"
                break
            except:
                pass
    
    # If no Thai font found, try to use THSarabun from uploads folder
    custom_font_path = os.path.join(app.config["UPLOAD_FOLDER"], "THSarabunNew.ttf")
    if thai_font == "Helvetica" and os.path.exists(custom_font_path):
        try:
            pdfmetrics.registerFont(TTFont("ThaiFont", custom_font_path))
            thai_font = "ThaiFont"
            thai_font_italic = "ThaiFont"
        except:
            pass
    
    # Colors
    primary_color = HexColor("#667eea")
    dark_color = HexColor("#1e293b")
    muted_color = HexColor("#64748b")
    bg_color = HexColor("#f8fafc")
    accent_color = HexColor("#10b981")
    
    # Adjust font size for Thai fonts (they're usually larger)
    base_size = 14 if thai_font != "Helvetica" else 12
    title_size = 24 if thai_font != "Helvetica" else 22
    
    def draw_bullet(x, y, text, max_width, font_size=None):
        # Draw a bullet point and return new y position
        if font_size is None:
            font_size = base_size
        if y < 2*cm:
            return y
        c.setFillColor(primary_color)
        c.circle(x, y + 0.12*cm, 0.1*cm, fill=1, stroke=0)
        c.setFillColor(dark_color)
        c.setFont(thai_font, font_size)
        wrapped = textwrap.wrap(str(text), width=int(max_width / 6))
        for line in wrapped[:3]:
            c.drawString(x + 0.5*cm, y, line)
            y -= 0.55*cm
        return y - 0.15*cm
    
    def extract_content_from_slide(slide):
        # Extract displayable content from any slide type
        slide_type = slide.get("type", "")
        items = []
        
        # Get content from various possible keys
        content = slide.get("content", [])
        if isinstance(content, str):
            items.append(content)
        elif isinstance(content, list):
            for ct in content:
                if isinstance(ct, str):
                    items.append(ct)
                elif isinstance(ct, dict):
                    en = ct.get("en") or ct.get("text") or ""
                    th = ct.get("th") or ct.get("meaning") or ""
                    if en:
                        items.append(f"{en}" + (f" ({th})" if th else ""))
        
        # Objectives
        objectives = slide.get("objectives", [])
        if isinstance(objectives, list):
            for obj in objectives:
                if isinstance(obj, str):
                    items.append(f"• {obj}")
        
        # Vocabulary
        vocabulary = slide.get("vocabulary", []) or slide.get("items", [])
        if isinstance(vocabulary, list) and slide_type in ["vocabulary", ""]:
            for v in vocabulary[:8]:
                if isinstance(v, dict):
                    word = v.get("word", "")
                    meaning = v.get("meaning", "") or v.get("th", "")
                    example = v.get("example", "") or v.get("example_en", "")
                    if word:
                        line = f"• {word}"
                        if meaning:
                            line += f" - {meaning}"
                        items.append(line)
                        if example:
                            items.append(f"  Ex: {example}")
        
        # Examples (en/th format)
        examples = slide.get("examples", [])
        if isinstance(examples, list):
            for ex in examples[:6]:
                if isinstance(ex, dict):
                    en = ex.get("en", "")
                    th = ex.get("th", "")
                    if en:
                        items.append(f"• {en}" + (f" ({th})" if th else ""))
                elif isinstance(ex, str):
                    items.append(f"• {ex}")
        
        # Highlights (for concept slides)
        highlights = slide.get("highlights", [])
        if isinstance(highlights, list):
            for hl in highlights[:5]:
                if isinstance(hl, dict):
                    label = hl.get("label", "")
                    note = hl.get("note", "")
                    if label:
                        items.append(f"• {label}: {note}" if note else f"• {label}")
        
        # Pattern/Structure (for concept slides)
        pattern = slide.get("pattern") or slide.get("structure", "")
        if pattern and isinstance(pattern, str):
            items.insert(0, f"📝 {pattern}")
        
        # Prompt (for hook slides)
        prompt = slide.get("prompt", "")
        if prompt and isinstance(prompt, str):
            items.insert(0, prompt)
        
        # Keywords
        keywords = slide.get("keywords", [])
        if isinstance(keywords, list) and keywords:
            items.append(f"Keywords: {', '.join(str(k) for k in keywords)}")
        
        # Dialogue lines
        lines = slide.get("lines", [])
        if isinstance(lines, list):
            for line in lines[:8]:
                if isinstance(line, dict):
                    speaker = line.get("speaker", "")
                    text = line.get("en") or line.get("text", "")
                    if speaker and text:
                        items.append(f"{speaker}: {text}")
                elif isinstance(line, str):
                    items.append(line)
        
        # Scenario
        scenario = slide.get("scenario", "")
        if scenario and isinstance(scenario, str):
            items.insert(0, f"🎭 {scenario}")
        
        # Guided practice items
        practice_items = slide.get("items", [])
        if isinstance(practice_items, list) and slide_type == "guided_practice":
            for pi in practice_items[:4]:
                if isinstance(pi, dict):
                    q = pi.get("q") or pi.get("question", "")
                    if q:
                        items.append(f"Q: {q}")
                    choices = pi.get("choices", [])
                    if choices:
                        items.append(f"   A) {choices[0] if len(choices) > 0 else ''}")
                        items.append(f"   B) {choices[1] if len(choices) > 1 else ''}")
                        items.append(f"   C) {choices[2] if len(choices) > 2 else ''}")
                        items.append(f"   D) {choices[3] if len(choices) > 3 else ''}")
        
        # Common mistakes
        mistakes = slide.get("common_mistakes", [])
        if isinstance(mistakes, list) and mistakes:
            items.append("")
            items.append("⚠️ Common mistakes:")
            for m in mistakes[:3]:
                items.append(f"  • {m}")
        
        # Bullets (generic)
        bullets = slide.get("bullets", [])
        if isinstance(bullets, list):
            for b in bullets:
                if isinstance(b, str):
                    items.append(f"• {b}")
        
        return items
    
    for i, slide in enumerate(slides):
        slide_title = slide.get("title", f"Slide {i+1}")
        slide_type = slide.get("type", "")
        image_url = slide.get("image_url", "")
        
        # Background
        c.setFillColor(bg_color)
        c.rect(0, 0, w, h, fill=1, stroke=0)
        
        # Header bar
        c.setFillColor(primary_color)
        c.rect(0, h - 2.5*cm, w, 2.5*cm, fill=1, stroke=0)
        
        # Slide number
        c.setFillColor(HexColor("#ffffff"))
        c.setFont("Helvetica", 10)
        c.drawRightString(w - 1*cm, h - 1.5*cm, f"{i+1} / {len(slides)}")
        
        # Slide type badge
        if slide_type:
            c.setFont("Helvetica", 8)
            c.drawRightString(w - 1*cm, h - 2*cm, f"[{slide_type}]")
        
        # Title
        c.setFillColor(HexColor("#ffffff"))
        c.setFont(thai_font_bold, title_size)
        display_title = slide_title[:55] + "..." if len(slide_title) > 55 else slide_title
        c.drawString(1.5*cm, h - 1.7*cm, display_title)
        
        y = h - 4*cm
        content_width = w - 3*cm
        
        # Check if there's an image
        img_x = None
        if image_url and not image_url.startswith("data:"):
            content_width = w * 0.55
            img_x = w * 0.58
        
        # Extract and draw content
        content_items = extract_content_from_slide(slide)
        
        c.setFillColor(dark_color)
        c.setFont(thai_font, base_size)
        
        for item in content_items:
            if y < 2*cm:
                break
            
            item_str = str(item).strip()
            if not item_str:
                y -= 0.3*cm
                continue
            
            # Check for special formatting
            if item_str.startswith("•"):
                y = draw_bullet(1.5*cm, y, item_str[1:].strip(), content_width, base_size)
            elif item_str.startswith("  •"):
                y = draw_bullet(2.2*cm, y, item_str[3:].strip(), content_width - 0.7*cm, base_size - 1)
            elif item_str.startswith("📝") or item_str.startswith("🎭") or item_str.startswith("⚠️"):
                c.setFont(thai_font_bold, base_size + 1)
                c.setFillColor(accent_color)
                wrapped = textwrap.wrap(item_str, width=int(content_width / 7))
                for line in wrapped[:2]:
                    c.drawString(1.5*cm, y, line)
                    y -= 0.6*cm
                c.setFillColor(dark_color)
                c.setFont(thai_font, base_size)
                y -= 0.2*cm
            elif item_str.startswith("Q:"):
                c.setFont(thai_font_bold, base_size)
                wrapped = textwrap.wrap(item_str, width=int(content_width / 7))
                for line in wrapped[:2]:
                    c.drawString(1.5*cm, y, line)
                    y -= 0.55*cm
                c.setFont(thai_font, base_size)
            elif item_str.startswith("   "):
                c.setFont(thai_font, base_size - 1)
                c.drawString(2*cm, y, item_str.strip())
                y -= 0.5*cm
                c.setFont(thai_font, base_size)
            elif item_str.startswith("Ex:") or item_str.startswith("  Ex:"):
                c.setFont(thai_font_italic, base_size - 1)
                c.setFillColor(muted_color)
                wrapped = textwrap.wrap(item_str, width=int(content_width / 6.5))
                for line in wrapped[:2]:
                    c.drawString(2*cm, y, line)
                    y -= 0.5*cm
                c.setFillColor(dark_color)
                c.setFont(thai_font, base_size)
            elif ":" in item_str and not item_str.startswith("Keywords"):
                parts = item_str.split(":", 1)
                c.setFont(thai_font_bold, base_size - 1)
                c.drawString(1.5*cm, y, parts[0] + ":")
                c.setFont(thai_font, base_size - 1)
                if len(parts) > 1:
                    wrapped = textwrap.wrap(parts[1].strip(), width=int(content_width / 7))
                    first = True
                    for line in wrapped[:2]:
                        if first:
                            c.drawString(1.5*cm + c.stringWidth(parts[0] + ": ", thai_font_bold, base_size - 1), y, line)
                            first = False
                        else:
                            c.drawString(2*cm, y, line)
                        y -= 0.55*cm
                else:
                    y -= 0.55*cm
                c.setFont(thai_font, base_size)
            else:
                wrapped = textwrap.wrap(item_str, width=int(content_width / 7))
                for line in wrapped[:3]:
                    c.drawString(1.5*cm, y, line)
                    y -= 0.55*cm
                y -= 0.1*cm
        
        # Image
        if image_url and img_x:
            try:
                img_path = None
                
                if image_url.startswith("/uploads/"):
                    img_path = os.path.join(app.config["UPLOAD_FOLDER"], image_url.split("/uploads/")[-1])
                elif image_url.startswith("http"):
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as tmp:
                        urllib.request.urlretrieve(image_url, tmp.name)
                        img_path = tmp.name
                
                if img_path and os.path.exists(img_path):
                    img_max_w = w * 0.38
                    img_max_h = h - 5*cm
                    
                    from reportlab.lib.utils import ImageReader
                    img = ImageReader(img_path)
                    iw, ih = img.getSize()
                    
                    scale = min(img_max_w / iw, img_max_h / ih)
                    draw_w = iw * scale
                    draw_h = ih * scale
                    
                    img_y = (h - 2.5*cm - draw_h) / 2
                    
                    c.drawImage(img_path, img_x, img_y, width=draw_w, height=draw_h, preserveAspectRatio=True)
            except Exception as e:
                print(f"Error loading image: {e}")
        
        # Footer
        c.setFillColor(muted_color)
        c.setFont(thai_font, 10)
        c.drawString(1.5*cm, 0.8*cm, title)
        
        c.showPage()
    
    c.save()
    return buf.getvalue()


# ==============================================================================
# Game
# ==============================================================================
@app.route("/topic/<int:topic_id>/game")
@login_required
def game(topic_id):
    topic = _get_topic_or_404(topic_id)
    last_session = GameSession.get_latest_by_topic_and_user(topic_id, session["user_id"])
    return render_template("game.html", topic=topic, last_session=last_session)

@app.route("/api/game/<int:topic_id>/sets")
@login_required
def api_game_sets(topic_id):
    _get_topic_or_404(topic_id)
    sets_data = {}
    for set_no in range(1, 4):
        questions = GameQuestion.get_by_topic_and_set(topic_id, set_no)
        if questions:
            sets_data[str(set_no)] = [{"id": q["id"], "tile_no": q["tile_no"], "question": q["question"], "answer": q["answer"], "points": q["points"]} for q in questions]
    return jsonify(sets_data)

@app.route("/api/game/<int:topic_id>/sessions", methods=["GET", "POST"])
@login_required
def api_game_sessions(topic_id):
    _get_topic_or_404(topic_id)
    if request.method == "GET":
        return jsonify({"ok": True, "sessions": GameSession.get_by_topic(topic_id)})
    data = request.get_json(silent=True) or {}
    sess = GameSession.create(topic_id, session["user_id"], data.get("title") or "Session", json.dumps(data.get("settings") or {}), json.dumps(data.get("state") or {}))
    return jsonify({"ok": True, "session": sess})

@app.route("/api/game/session/<int:session_id>")
@login_required
def api_game_session_get(session_id):
    sess = GameSession.get_by_id(session_id)
    return jsonify({"ok": True, "session": sess}) if sess else _json_error("Not found", 404)

@app.route("/api/game/session/<int:session_id>/save", methods=["POST"])
@login_required
def api_game_session_save(session_id):
    sess = GameSession.get_by_id(session_id)
    if not sess: return _json_error("Not found", 404)
    data = request.get_json(silent=True) or {}
    GameSession.update(session_id, data.get("title") or sess["title"], json.dumps(data.get("settings") or {}), json.dumps(data.get("state") or {}))
    return jsonify({"ok": True})


# ==============================================================================
# Memory Match Game
# ==============================================================================
@app.route("/topic/<int:topic_id>/game/memory")
@login_required
def game_memory(topic_id):
    topic = _get_topic_or_404(topic_id)
    
    # Get vocabulary from slides
    vocabulary = []
    if topic.get("slides_json"):
        try:
            obj = json.loads(topic["slides_json"])
            slides = obj.get("slides", obj) if isinstance(obj, dict) else obj
            for slide in slides:
                if slide.get("type") == "vocabulary" and slide.get("vocabulary"):
                    for v in slide["vocabulary"]:
                        if v.get("word") and v.get("meaning"):
                            vocabulary.append({"word": v["word"], "meaning": v["meaning"]})
        except:
            pass
    
    # Get game questions as fallback
    questions = []
    for set_no in range(1, 4):
        qs = GameQuestion.get_by_topic_and_set(topic_id, set_no)
        for q in qs:
            questions.append({"question": q["question"], "answer": q["answer"]})
    
    game_data = {"vocabulary": vocabulary, "questions": questions}
    return render_template("game_memory.html", topic=topic, game_data=game_data)


# ==============================================================================
# Millionaire Game
# ==============================================================================
@app.route("/topic/<int:topic_id>/game/millionaire")
@login_required
def game_millionaire(topic_id):
    topic = _get_topic_or_404(topic_id)
    
    # Get practice questions (MCQ)
    questions = _normalize_practice_questions(PracticeQuestion.get_by_topic(topic_id))
    
    return render_template("game_millionaire.html", topic=topic, questions=questions)


# ==============================================================================
# Sentence Builder: Helpers & Logic
# ==============================================================================
def _topic_slides_obj(topic):
    # Parse topic['slides_json'] into dict. Always returns dict.
    try:
        raw = topic.get("slides_json") or ""
        obj = json.loads(raw) if raw else {}
        if isinstance(obj, list):
            # legacy list of slides
            return {"slides": obj}
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass
    return {"slides": []}

def _topic_get_sentence_builder_custom(topic):
    obj = _topic_slides_obj(topic)
    items = obj.get("sentence_builder_custom") or []
    out = []
    if isinstance(items, list):
        for it in items:
            if isinstance(it, dict):
                th = (it.get("th") or "").strip()
                en = (it.get("en") or "").strip()
                if th or en:
                    out.append({"th": th, "en": en})
    return out

def _topic_save_sentence_builder_custom(topic_id, items):
    topic = Topic.get_by_id(topic_id)
    if not topic:
        return False
    obj = _topic_slides_obj(topic)
    cleaned = []
    if isinstance(items, list):
        for it in items:
            if not isinstance(it, dict):
                continue
            th = (it.get("th") or "").strip()
            en = (it.get("en") or "").strip()
            if not th and not en:
                continue
            cleaned.append({"th": th[:500], "en": en[:500]})
    obj["sentence_builder_custom"] = cleaned
    Topic.update(topic_id, topic["name"], topic.get("description") or "", json.dumps(obj, ensure_ascii=False), topic.get("pdf_file"))
    return True

# ==============================================================================
# Sentence Builder Game
# ==============================================================================

# API: Save custom sentences
@app.route("/api/topic/<int:topic_id>/sentence-builder/custom", methods=["GET", "POST"])
@login_required
def api_sentence_builder_custom(topic_id):
    topic = _get_topic_or_404(topic_id)
    if request.method == "GET":
        return jsonify({"ok": True, "items": _topic_get_sentence_builder_custom(topic)})
    data = request.get_json(silent=True) or {}
    items = data.get("items") or []
    ok = _topic_save_sentence_builder_custom(topic_id, items)
    return jsonify({"ok": bool(ok), "items": _topic_get_sentence_builder_custom(Topic.get_by_id(topic_id))})

# Main View
@app.route("/topic/<int:topic_id>/game/sentence-builder")
@login_required
def game_sentence_builder(topic_id):
    # Game: เรียงประโยคภาษาอังกฤษจากประโยคภาษาไทย
    topic = _get_topic_or_404(topic_id)
    game_data = _get_practice_data_from_slides(topic)
    game_data = _sentence_builder_enrich_game_data_with_th(topic, game_data)
    
    # Get students from classroom if linked
    students = []
    # Try to get students from classroom assignments
    conn = get_db()
    c = conn.cursor()
    # ดึงข้อมูลจาก classroom_students โดยตรง และ join กับ assignments
    c.execute("""
        SELECT DISTINCT cs.student_name 
        FROM classroom_students cs
        JOIN assignments a ON a.classroom_id = cs.classroom_id
        WHERE a.topic_id = ?
        ORDER BY cs.student_no, cs.student_name
    """, (topic_id,))
    rows = c.fetchall()
    conn.close()
    
    students = [r["student_name"] for r in rows] if rows else []
    
    return render_template("game_sentence_builder.html", topic=topic, game_data=game_data, students=students)


# ==============================================================================
# Practice Helpers
# ==============================================================================
def _normalize_practice_questions(rows):
    out = []
    for r in rows:
        q = dict(r)
        prompt, choices = "", []
        raw = q.get("question") or ""
        try:
            obj = json.loads(raw)
            if isinstance(obj, dict):
                prompt = (obj.get("prompt") or "").strip()
                choices = [str(x) for x in (obj.get("choices") or [])]
            else: prompt = str(obj)
        except: prompt = str(raw)
        out.append({"id": q.get("id"), "prompt": prompt, "choices": choices, "correct_answer": q.get("correct_answer") or ""})
    return out

def _build_practice_pdf(topic_title, questions, include_answers=False):
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    mx, y, lh, mw = 2*cm, h-2*cm, 14, w-4*cm
    def draw(lines, y0):
        y = y0
        for ln in lines:
            if y < 2*cm: c.showPage(); y = h-2*cm
            c.drawString(mx, y, ln); y -= lh
        return y
    c.setFont("Helvetica-Bold", 16)
    y = draw([f"Practice: {topic_title}"], y)
    c.setFont("Helvetica", 11)
    y = draw(["Name: ________________________   Class: __________", ""], y)
    for i, q in enumerate(questions, 1):
        c.setFont("Helvetica-Bold", 12)
        y = draw(simpleSplit(f"{i}. {q.get('prompt','')}", "Helvetica-Bold", 12, mw), y)
        c.setFont("Helvetica", 11)
        ch = q.get("choices") or []
        if len(ch) == 4:
            for lab, cv in zip(["A","B","C","D"], ch):
                y = draw(simpleSplit(f"   ({lab}) {cv}", "Helvetica", 11, mw), y)
        if include_answers: y = draw([f"   Answer: {q.get('correct_answer','')}"], y)
        y = draw([""], y)
    c.showPage(); c.save()
    return buf.getvalue()


# ==============================================================================
# Practice
# ==============================================================================
@app.route("/topic/<int:topic_id>/practice")
@login_required
def practice(topic_id):
    topic = _get_topic_or_404(topic_id)
    questions = _normalize_practice_questions(PracticeQuestion.get_by_topic(topic_id))
    link = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "mcq")
    student_url = (request.url_root.rstrip("/") + url_for("public_practice", token=link["token"])) if link else None
    return render_template("practice.html", topic=topic, questions=questions, student_url=student_url)


def _get_practice_data_from_slides(topic):
    # Extract vocabulary, examples, dialogues from slides for practice activities
    data = {"vocabulary": [], "examples": [], "dialogues": [], "questions": [], "mcq_questions": []}
    
    # From slides
    if topic.get("slides_json"):
        try:
            obj = json.loads(topic["slides_json"])
            slides = obj.get("slides", obj) if isinstance(obj, dict) else obj
            for slide in slides:
                slide_type = slide.get("type", "")
                
                # Vocabulary
                if slide_type == "vocabulary" and slide.get("vocabulary"):
                    for v in slide["vocabulary"]:
                        if v.get("word") and v.get("meaning"):
                            data["vocabulary"].append({
                                "word": v["word"],
                                "meaning": v["meaning"],
                                "example": v.get("example", "")
                            })
                
                # Examples
                if slide.get("examples"):
                    for ex in slide["examples"]:
                        if isinstance(ex, dict) and ex.get("en"):
                            data["examples"].append({"en": ex["en"], "th": ex.get("th", "")})
                        elif isinstance(ex, str):
                            data["examples"].append({"en": ex, "th": ""})
                
                # Dialogues
                if slide_type == "dialogue" and slide.get("lines"):
                    for line in slide["lines"]:
                        if isinstance(line, dict) and line.get("text"):
                            data["dialogues"].append({"speaker": line.get("speaker", ""), "text": line["text"]})
        except:
            pass
    
    # From game questions
    for set_no in range(1, 4):
        qs = GameQuestion.get_by_topic_and_set(topic["id"], set_no)
        for q in qs:
            data["questions"].append({"question": q["question"], "answer": q["answer"]})
    
    # From MCQ practice questions
    mcq_rows = _normalize_practice_questions(PracticeQuestion.get_by_topic(topic["id"]))
    data["mcq_questions"] = mcq_rows
    
    return data


# ------------------------------------------------------------------------------
# Sentence Builder helpers (Auto-generate Thai translations from slides if missing)
# ------------------------------------------------------------------------------
def _extract_first_json_array(s: str):
    # Best-effort: extract first JSON array from a string.
    if not s:
        return None
    s = s.strip()
    # already json
    if s.startswith('[') and s.endswith(']'):
        try:
            return json.loads(s)
        except Exception:
            pass
    # find first [...]
    start = s.find('[')
    if start == -1:
        return None
    depth = 0
    for i in range(start, len(s)):
        ch = s[i]
        if ch == '[':
            depth += 1
        elif ch == ']':
            depth -= 1
            if depth == 0:
                chunk = s[start:i+1]
                try:
                    return json.loads(chunk)
                except Exception:
                    return None
    return None

def _ai_translate_en_to_th(sentences, model="gpt-4o-mini"):
    # Translate a list of short English sentences to Thai (returns list of dicts: {en, th}).
    api_key = os.environ.get("OPENAI_API_KEY") or os.environ.get("OPENAI_APIKEY") or os.environ.get("OPENAI_KEY")
    if not api_key:
        return []
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key)
    except Exception:
        return []

    # Keep it strict JSON to parse reliably
    sys = (
        "You translate English teaching examples into natural Thai. "
        "Return ONLY valid JSON array. No markdown. No extra text."
    )
    user = {
        "task": "translate_en_to_th",
        "rules": [
            "Keep meaning faithful and natural for Thai students.",
            "Do not add explanations.",
            "Do not number items.",
            "Return format: [{\"en\":...,\"th\":...}, ...] in same order."
        ],
        "sentences": sentences[:40]  # safety cap
    }
    try:
        # Compatible with OpenAI Python SDK 1.x
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": sys},
                {"role": "user", "content": json.dumps(user, ensure_ascii=False)}
            ],
            temperature=0.2,
        )
        content = (resp.choices[0].message.content or "").strip()
    except Exception:
        # Fallback: newer responses API if present
        try:
            resp = client.responses.create(
                model=model,
                input=[
                    {"role": "system", "content": sys},
                    {"role": "user", "content": json.dumps(user, ensure_ascii=False)}
                ],
                temperature=0.2,
            )
            content = (resp.output_text or "").strip()
        except Exception:
            return []

    arr = _extract_first_json_array(content)
    if not isinstance(arr, list):
        return []
    out = []
    for it in arr:
        if isinstance(it, dict) and it.get("en") and it.get("th"):
            out.append({"en": str(it["en"]).strip(), "th": str(it["th"]).strip()})
    return out

def _sentence_builder_enrich_game_data_with_th(topic, game_data):
    # Ensure examples contain Thai prompts for Sentence Builder.
    if not game_data:
        return game_data

    examples = (game_data.get("examples") or [])
    if not isinstance(examples, list) or not examples:
        return game_data

    # Collect EN sentences that are missing Thai (or Thai is not actually Thai characters)
    need_en = []
    for ex in examples:
        if not isinstance(ex, dict):
            continue
        en = (ex.get("en") or "").strip()
        th = (ex.get("th") or "").strip()
        has_thai = bool(th and re.search(r"[ก-๙]", th))
        if en and not has_thai:
            if en not in need_en:
                need_en.append(en)

    # Nothing to translate
    if not need_en:
        return game_data

    translated = _ai_translate_en_to_th(need_en)
    if not translated:
        return game_data

    mapping = {str(t.get("en") or "").strip(): str(t.get("th") or "").strip()
               for t in translated if isinstance(t, dict)}

    new_examples = []
    for ex in examples:
        if isinstance(ex, dict) and ex.get("en"):
            en = str(ex.get("en") or "").strip()
            th = str(ex.get("th") or "").strip()
            has_thai = bool(th and re.search(r"[ก-๙]", th))
            if (not has_thai) and en in mapping and mapping[en]:
                th = mapping[en]
            new_examples.append({"en": en, "th": th})
        else:
            new_examples.append(ex)

    game_data["examples"] = new_examples
    return game_data


@app.route("/api/practice/<int:topic_id>/fill-blanks/link", methods=["POST"])
@login_required
def api_fill_blanks_create_link(topic_id):
    _get_topic_or_404(topic_id)
    # ค้นหา link ที่เป็น fill โดยเฉพาะ
    old = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "fill")
    if not old:
        link = PracticeLink.create(topic_id, session["user_id"], secrets.token_urlsafe(12), "fill")
    else:
        link = old
    return jsonify({"url": request.url_root.rstrip("/") + url_for("public_fill_blanks", token=link["token"])})




@app.route("/topic/<int:topic_id>/practice/fill-blanks")
@login_required
def practice_fill_blanks(topic_id):
    topic = _get_topic_or_404(topic_id)
    practice_data = _get_practice_data_from_slides(topic)
    link = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "fill")
    student_url = None
    if link:
        student_url = request.url_root.rstrip("/") + url_for("public_fill_blanks", token=link["token"])
    return render_template("practice_fill_blanks.html", topic=topic, practice_data=practice_data, student_url=student_url)


@app.route("/topic/<int:topic_id>/practice/fill-blanks/scores")
@login_required
def practice_fill_blanks_scores(topic_id):
    topic = _get_topic_or_404(topic_id)
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT ps.* FROM practice_submissions ps JOIN practice_links pl ON ps.link_id=pl.id WHERE pl.topic_id=? AND pl.practice_type='fill' ORDER BY ps.id DESC LIMIT 500", (topic_id,))
    submissions = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template("practice_scores.html", topic=topic, submissions=submissions, practice_type="Fill in the Blanks")


@app.route("/p/fill/<token>")
def public_fill_blanks(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link["is_active"]:
        return "ลิงก์ไม่ถูกต้องหรือหมดอายุ", 404
    topic = Topic.get_by_id(link["topic_id"])
    if not topic:
        return "Topic not found", 404
    practice_data = _get_practice_data_from_slides(topic)
    
    # Get classrooms of the teacher who created the link
    classrooms = Classroom.get_by_owner(link["created_by"]) if link.get("created_by") else []
    
    return render_template("practice_fill_blanks_public.html", topic=topic, practice_data=practice_data, token=token, classrooms=classrooms)


@app.route("/api/public/fill/<token>/submit", methods=["POST"])
def api_public_fill_blanks_submit(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link["is_active"]:
        return _json_error("Invalid link", 404)
    data = request.get_json() or {}
    student_name = (data.get("student_name") or data.get("name") or "Anonymous").strip()[:100]
    student_no = (data.get("student_no") or "").strip()[:20]
    classroom = (data.get("classroom") or "").strip()[:30]
    score = int(data.get("score", 0))
    total = int(data.get("total", 0))
    pct = (score/total*100) if total else 0
    PracticeSubmission.create(link["id"], student_name, student_no, classroom, json.dumps(data.get("answers", {})), score, total, pct)
    return jsonify({"ok": True, "score": score, "total": total, "percentage": pct})


@app.route("/topic/<int:topic_id>/practice/unscramble")
@login_required
def practice_unscramble(topic_id):
    topic = _get_topic_or_404(topic_id)
    practice_data = _get_practice_data_from_slides(topic)
    link = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "unscramble")
    student_url = None
    if link:
        student_url = request.url_root.rstrip("/") + url_for("public_unscramble", token=link["token"])
    return render_template("practice_unscramble.html", topic=topic, practice_data=practice_data, student_url=student_url)


@app.route("/api/practice/<int:topic_id>/unscramble/link", methods=["POST"])
@login_required
def api_unscramble_create_link(topic_id):
    _get_topic_or_404(topic_id)
    # ค้นหา link ที่เป็น unscramble โดยเฉพาะ
    old = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "unscramble")
    if not old:
        link = PracticeLink.create(topic_id, session["user_id"], secrets.token_urlsafe(12), "unscramble")
    else:
        link = old
    return jsonify({"url": request.url_root.rstrip("/") + url_for("public_unscramble", token=link["token"])})


@app.route("/topic/<int:topic_id>/practice/unscramble/scores")
@login_required
def practice_unscramble_scores(topic_id):
    topic = _get_topic_or_404(topic_id)
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT ps.* FROM practice_submissions ps JOIN practice_links pl ON ps.link_id=pl.id WHERE pl.topic_id=? AND pl.practice_type='unscramble' ORDER BY ps.id DESC LIMIT 500", (topic_id,))
    submissions = [dict(r) for r in c.fetchall()]
    conn.close()
    return render_template("practice_scores.html", topic=topic, submissions=submissions, practice_type="Sentence Unscramble")


@app.route("/p/unscramble/<token>")
def public_unscramble(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link["is_active"]:
        return "ลิงก์ไม่ถูกต้องหรือหมดอายุ", 404
    topic = Topic.get_by_id(link["topic_id"])
    if not topic:
        return "Topic not found", 404
    practice_data = _get_practice_data_from_slides(topic)
    
    # Get classrooms of the teacher who created the link
    classrooms = Classroom.get_by_owner(link["created_by"]) if link.get("created_by") else []
    
    return render_template("practice_unscramble_public.html", topic=topic, practice_data=practice_data, token=token, classrooms=classrooms)


@app.route("/api/public/unscramble/<token>/submit", methods=["POST"])
def api_public_unscramble_submit(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link["is_active"]:
        return _json_error("Invalid link", 404)
    data = request.get_json() or {}
    student_name = (data.get("student_name") or data.get("name") or "Anonymous").strip()[:100]
    student_no = (data.get("student_no") or "").strip()[:20]
    classroom = (data.get("classroom") or "").strip()[:30]
    score = int(data.get("score", 0))
    total = int(data.get("total", 0))
    pct = (score/total*100) if total else 0
    PracticeSubmission.create(link["id"], student_name, student_no, classroom, json.dumps(data.get("answers", {})), score, total, pct)
    return jsonify({"ok": True, "score": score, "total": total, "percentage": pct})

@app.route("/api/practice/<int:topic_id>/submit", methods=["POST"])
@login_required
def api_practice_submit(topic_id):
    _get_topic_or_404(topic_id)
    data = request.get_json() or {}
    answers = data.get("answers", {})
    questions = _normalize_practice_questions(PracticeQuestion.get_by_topic(topic_id))
    score, total, feedback = 0, len(questions), {}
    for q in questions:
        qid = str(q["id"])
        ua = (answers.get(qid, "") or "").strip().lower()
        ca = (q.get("correct_answer") or "").strip().lower()
        correct = ua == ca
        if correct: score += 1
        feedback[qid] = {"is_correct": correct, "user_answer": answers.get(qid, ""), "correct_answer": q.get("correct_answer")}
    pct = (score/total*100) if total else 0
    AttemptHistory.create(session["user_id"], topic_id, score, total, pct)
    return jsonify({"score": score, "total": total, "percentage": pct, "feedback": feedback})

@app.route("/api/practice/<int:topic_id>/link", methods=["POST"])
@login_required
def api_practice_create_link(topic_id):
    _get_topic_or_404(topic_id)
    # ค้นหา link ที่เป็น mcq โดยเฉพาะ
    old = PracticeLink.get_by_topic_user_and_type(topic_id, session["user_id"], "mcq")
    if not old:
        link = PracticeLink.create(topic_id, session["user_id"], secrets.token_urlsafe(12), "mcq")
    else:
        link = old
    return jsonify({"url": request.url_root.rstrip("/") + url_for("public_practice", token=link["token"])})

@app.route("/topic/<int:topic_id>/practice/pdf")
@login_required
def practice_pdf(topic_id):
    topic = _get_topic_or_404(topic_id)
    include_answers = request.args.get("answers") == "1"
    pdf = _build_practice_pdf(topic["name"], _normalize_practice_questions(PracticeQuestion.get_by_topic(topic_id)), include_answers)
    return (pdf, 200, {"Content-Type": "application/pdf", "Content-Disposition": f"attachment; filename=practice_{topic_id}.pdf"})

@app.route("/topic/<int:topic_id>/practice/scores")
@login_required
def practice_scores(topic_id):
    topic = _get_topic_or_404(topic_id)
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT ps.* FROM practice_submissions ps JOIN practice_links pl ON ps.link_id=pl.id WHERE pl.topic_id=? AND pl.practice_type='mcq' ORDER BY ps.id DESC LIMIT 1000", (topic_id,))
    submissions = [dict(r) for r in c.fetchall()]
    conn.close()
    classrooms = sorted(set(s.get("classroom") or "" for s in submissions if s.get("classroom")))
    return render_template("practice_scores.html", topic=topic, submissions=submissions, classrooms=classrooms)

@app.route("/topic/<int:topic_id>/practice/scores/csv")
@login_required
def practice_scores_csv(topic_id):
    topic = _get_topic_or_404(topic_id)
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT ps.* FROM practice_submissions ps JOIN practice_links pl ON ps.link_id=pl.id WHERE pl.topic_id=? AND pl.practice_type='mcq' ORDER BY ps.classroom,ps.student_no", (topic_id,))
    rows = c.fetchall()
    conn.close()
    out = StringIO()
    w = csv.writer(out)
    w.writerow(["#", "Name", "No", "Class", "Score", "Total", "%", "Time"])
    for i, r in enumerate(rows, 1):
        w.writerow([i, r["student_name"], r["student_no"] or "", r["classroom"] or "", r["score"], r["total"], f"{r['percentage']:.0f}%", r["created_at"]])
    return Response(out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename=scores_{topic_id}.csv"})

@app.route("/topic/<int:topic_id>/practice/scores/excel")
@login_required
def practice_scores_excel(topic_id):
    topic = _get_topic_or_404(topic_id)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except: return redirect(url_for("practice_scores_csv", topic_id=topic_id))
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT ps.* FROM practice_submissions ps JOIN practice_links pl ON ps.link_id=pl.id WHERE pl.topic_id=? AND pl.practice_type='mcq' ORDER BY ps.classroom,ps.student_no", (topic_id,))
    rows = c.fetchall()
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"
    hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="667eea")
    bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Practice Scores: {topic['name']}"
    ws['A1'].font = Font(bold=True, size=14)
    for col, h in enumerate(["#", "Name", "No", "Class", "Score", "Total", "%", "Time"], 1):
        cell = ws.cell(3, col, h)
        cell.font, cell.fill, cell.border = hf, hfill, bd
    for i, r in enumerate(rows, 1):
        for col, v in enumerate([i, r["student_name"], r["student_no"] or "", r["classroom"] or "", r["score"], r["total"], f"{r['percentage']:.0f}%", str(r["created_at"])[:19]], 1):
            cell = ws.cell(i+3, col, v)
            cell.border = bd
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=scores_{topic_id}.xlsx"})



# ==============================================================================
# Public Practice
# ==============================================================================
@app.route("/p/<token>")
def public_practice(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link.get("is_active"): return render_template("error.html", error_code=404, error_msg="ลิงก์หมดอายุ"), 404
    topic = Topic.get_by_id(link["topic_id"])
    if not topic: return render_template("error.html", error_code=404, error_msg="Topic not found"), 404
    
    # Get classrooms of the teacher who created the link
    classrooms = Classroom.get_by_owner(link["created_by"]) if link.get("created_by") else []
    
    return render_template("practice_public.html", topic=topic, questions=_normalize_practice_questions(PracticeQuestion.get_by_topic(topic["id"])), token=token, classrooms=classrooms)

@app.route("/api/p/<token>/submit", methods=["POST"])
def api_public_practice_submit(token):
    link = PracticeLink.get_by_token(token)
    if not link or not link.get("is_active"): return jsonify({"error": "Invalid link"}), 404
    data = request.get_json() or {}
    name = (data.get("student_name") or "").strip()
    if not name: return jsonify({"error": "Name required"}), 400
    questions = _normalize_practice_questions(PracticeQuestion.get_by_topic(link["topic_id"]))
    answers = data.get("answers", {})
    score, total, feedback = 0, len(questions), {}
    for q in questions:
        qid = str(q["id"])
        ua = (answers.get(qid, "") or "").strip().lower()
        ca = (q.get("correct_answer") or "").strip().lower()
        correct = ua == ca
        if correct: score += 1
        feedback[qid] = {"is_correct": correct, "user_answer": answers.get(qid, ""), "correct_answer": q.get("correct_answer")}
    pct = (score/total*100) if total else 0
    PracticeSubmission.create(link["id"], name, data.get("student_no") or "", data.get("classroom") or "", json.dumps({"answers": answers}), score, total, pct)
    return jsonify({"score": score, "total": total, "percentage": pct, "feedback": feedback})

# API to get students by classroom (for public practice)
@app.route("/api/public/classroom/<int:classroom_id>/students")
def api_public_classroom_students(classroom_id):
    students = ClassroomStudent.get_by_classroom(classroom_id)
    return jsonify([{"id": s["id"], "student_no": s.get("student_no") or "", "student_name": s.get("student_name") or ""} for s in students])

# API: Get all classrooms for current user (for dropdown)
@app.route("/api/classrooms")
@login_required
def api_get_classrooms():
    # Get all classrooms for current user
    user_id = session["user_id"]
    cls_list = Classroom.get_by_owner(user_id)
    return jsonify({
        "classrooms": [{"id": c["id"], "name": c["name"]} for c in cls_list]
    })


# API: Get students in a classroom (for importing to games)
@app.route("/api/classroom/<int:classroom_id>/students")
@login_required
def api_get_classroom_students(classroom_id):
    # Get students in a classroom
    # Verify ownership
    classroom = Classroom.get_by_id(classroom_id)
    if not classroom or classroom.get("owner_id") != session["user_id"]:
        return jsonify({"error": "Not found"}), 404
    
    students = ClassroomStudent.get_by_classroom(classroom_id)
    return jsonify({
        "students": [
            {"id": s["id"], "student_no": s.get("student_no") or "", "student_name": s.get("student_name") or ""}
            for s in students
        ]
    })

# ==============================================================================
# Classrooms
# ==============================================================================
@app.route("/classrooms")
@login_required
def classrooms():
    user_id = session["user_id"]
    cls_list = Classroom.get_by_owner(user_id)
    total_students = sum(c.get("student_count") or 0 for c in cls_list)
    assignments = Assignment.get_by_owner(user_id)
    # Add assignment count to each classroom
    for c in cls_list:
        c["assignment_count"] = len([a for a in assignments if a["classroom_id"] == c["id"]])
    return render_template("classrooms.html", classrooms=cls_list, total_students=total_students, total_assignments=len(assignments))

@app.route("/classrooms/create", methods=["POST"])
@login_required
def classroom_create():
    # ตรวจสอบ limit (Freemium)
    is_premium = is_premium_user(session["user_id"])
    can_create, msg = UsageLimits.can_create_classroom(session["user_id"], is_premium)

    if not can_create:
        flash(f"❌ {msg} - อัปเกรดเป็น Premium!", "error")
        return redirect(url_for("pricing"))

    name = (request.form.get("name") or "").strip()
    if not name:
        flash("กรุณาระบุชื่อห้อง", "error")
        return redirect(url_for("classrooms"))

    Classroom.create(
        session["user_id"],
        name,
        request.form.get("grade_level") or "",
        request.form.get("academic_year") or "",
        request.form.get("description") or ""
    )
    flash("✅ สร้างห้องเรียนแล้ว", "success")
    return redirect(url_for("classrooms"))

@app.route("/classroom/<int:classroom_id>")
@login_required
def classroom_detail(classroom_id):
    cls = Classroom.get_by_id(classroom_id)
    if not cls or cls["owner_id"] != session["user_id"]: abort(404)
    students = ClassroomStudent.get_by_classroom(classroom_id)
    assignments = Assignment.get_by_classroom(classroom_id)
    topics = Topic.get_by_owner(session["user_id"])
    
    # Get submission stats and scores for each assignment
    submission_stats = {}
    assignment_stats = {}
    scores_by_student = {s["id"]: {"assignments": {}, "total_score": 0, "total_possible": 0} for s in students}
    
    for a in assignments:
        status = Assignment.get_submissions_status(a["id"])
        submission_stats[a["id"]] = {"submitted": len(status["submitted"]), "not_submitted": len(status["not_submitted"])}
        
        # Calculate assignment average
        submissions = status.get("submissions") or []
        if submissions:
            avg = sum(s.get("percentage") or 0 for s in submissions) / len(submissions)
            assignment_stats[a["id"]] = {"avg": avg, "count": len(submissions)}
        else:
            assignment_stats[a["id"]] = {"avg": 0, "count": 0}
        
        # Map submissions to students
        for student in students:
            student_id = student["id"]
            student_name_lower = (student.get("student_name") or "").strip().lower()
            student_no = (student.get("student_no") or "").strip()
            
            # Find matching submission
            for sub in submissions:
                sub_name = (sub.get("student_name") or "").strip().lower()
                sub_no = (sub.get("student_no") or "").strip()
                if sub_name == student_name_lower or (sub_no and sub_no == student_no):
                    scores_by_student[student_id]["assignments"][a["id"]] = {
                        "score": sub.get("score", 0),
                        "total": sub.get("total", 0),
                        "percentage": sub.get("percentage", 0)
                    }
                    scores_by_student[student_id]["total_score"] += sub.get("score", 0)
                    scores_by_student[student_id]["total_possible"] += sub.get("total", 0)
                    break
    
    # Calculate class average
    class_avg = 0
    students_with_scores = [s for s in scores_by_student.values() if s["total_possible"] > 0]
    if students_with_scores:
        class_avg = sum((s["total_score"] / s["total_possible"] * 100) for s in students_with_scores) / len(students_with_scores)
    
    return render_template("classroom_detail.html", classroom=cls, students=students, assignments=assignments, topics=topics, submission_stats=submission_stats, scores_by_student=scores_by_student, assignment_stats=assignment_stats, class_avg=class_avg)

@app.route("/classroom/<int:classroom_id>/edit", methods=["POST"])
@login_required
def classroom_edit(classroom_id):
    cls = Classroom.get_by_id(classroom_id)
    if not cls or cls["owner_id"] != session["user_id"]: abort(404)
    Classroom.update(classroom_id, request.form.get("name") or cls["name"], request.form.get("grade_level") or "", request.form.get("academic_year") or "", request.form.get("description") or "")
    flash("บันทึกแล้ว", "success")
    return redirect(url_for("classrooms"))

@app.route("/classroom/<int:classroom_id>/delete", methods=["POST"])
@login_required
def classroom_delete(classroom_id):
    cls = Classroom.get_by_id(classroom_id)
    if not cls or cls["owner_id"] != session["user_id"]: abort(404)
    Classroom.delete(classroom_id)
    flash("ลบห้องเรียนแล้ว", "success")
    return redirect(url_for("classrooms"))

@app.route("/classroom/<int:classroom_id>/add-student", methods=["POST"])
@login_required
def classroom_add_student(classroom_id):
    cls = Classroom.get_by_id(classroom_id)
    if not cls or cls["owner_id"] != session["user_id"]: abort(404)
    name = (request.form.get("student_name") or "").strip()
    if name:
        ClassroomStudent.create(classroom_id, request.form.get("student_no") or "", name, request.form.get("nickname") or "")
        flash("เพิ่มนักเรียนแล้ว", "success")
    return redirect(url_for("classroom_detail", classroom_id=classroom_id))

@app.route("/classroom/<int:classroom_id>/import-students", methods=["POST"])
@login_required
def classroom_import_students(classroom_id):
    cls = Classroom.get_by_id(classroom_id)
    if not cls or cls["owner_id"] != session["user_id"]: abort(404)
    text = request.form.get("student_list") or ""
    students = []
    for line in text.strip().split("\n"):
        line = line.strip()
        if not line: continue
        parts = line.split("\t")
        if len(parts) >= 2:
            students.append({"student_no": parts[0].strip(), "student_name": parts[1].strip()})
        else:
            students.append({"student_no": "", "student_name": parts[0].strip()})
    count = ClassroomStudent.bulk_create(classroom_id, students)
    flash(f"Import {count} คนเรียบร้อย", "success")
    return redirect(url_for("classroom_detail", classroom_id=classroom_id))

@app.route("/classroom/student/<int:student_id>/edit", methods=["POST"])
@login_required
def classroom_student_edit(student_id):
    s = ClassroomStudent.get_by_id(student_id)
    if not s: abort(404)
    cls = Classroom.get_by_id(s["classroom_id"])
    if not cls or cls["owner_id"] != session["user_id"]: abort(404)
    ClassroomStudent.update(student_id, request.form.get("student_no") or "", request.form.get("student_name") or s["student_name"], request.form.get("nickname") or "")
    return redirect(url_for("classroom_detail", classroom_id=s["classroom_id"]))

@app.route("/classroom/student/<int:student_id>/delete", methods=["POST"])
@login_required
def classroom_student_delete(student_id):
    s = ClassroomStudent.get_by_id(student_id)
    if not s: abort(404)
    cls = Classroom.get_by_id(s["classroom_id"])
    if not cls or cls["owner_id"] != session["user_id"]: abort(404)
    classroom_id = s["classroom_id"]
    ClassroomStudent.delete(student_id)
    return redirect(url_for("classroom_detail", classroom_id=classroom_id))

@app.route("/classroom/<int:classroom_id>/assign", methods=["POST"])
@login_required
def classroom_assign(classroom_id):
    cls = Classroom.get_by_id(classroom_id)
    if not cls or cls["owner_id"] != session["user_id"]: abort(404)
    topic_id = int(request.form.get("topic_id") or 0)
    if not topic_id:
        flash("กรุณาเลือก Topic", "error")
        return redirect(url_for("classroom_detail", classroom_id=classroom_id))
    topic = Topic.get_by_id(topic_id)
    if not topic: abort(404)
    # Create practice link
    exercise_type = request.form.get("exercise_type", "mcq")
    link = PracticeLink.create(topic_id, session["user_id"], secrets.token_urlsafe(12), exercise_type)
    title = (request.form.get("title") or "").strip() or topic["name"]
    due_date = request.form.get("due_date") or None
    Assignment.create(classroom_id, topic_id, link["id"], title, request.form.get("description") or "", due_date, session["user_id"])
    flash("สั่งงานเรียบร้อย", "success")
    return redirect(url_for("classroom_detail", classroom_id=classroom_id))

@app.route("/assignment/<int:assignment_id>")
@login_required
def assignment_detail(assignment_id):
    a = Assignment.get_by_id(assignment_id)
    if not a: abort(404)
    cls = Classroom.get_by_id(a["classroom_id"])
    if not cls or cls["owner_id"] != session["user_id"]: abort(404)
    topic = Topic.get_by_id(a["topic_id"])
    status = Assignment.get_submissions_status(assignment_id)
    practice_link = PracticeLink.get_by_id(a.get("practice_link_id")) if a.get("practice_link_id") else None
    student_url = (request.url_root.rstrip("/") + url_for("public_practice", token=practice_link["token"])) if practice_link else None
    # Calculate average score
    avg = 0
    submissions = status.get("submissions") or []
    if submissions:
        avg = sum(s.get("percentage") or 0 for s in submissions) / len(submissions)
    return render_template("assignment_detail.html", assignment=a, classroom=cls, topic=topic, submitted=status["submitted"], not_submitted=status["not_submitted"], total_students=status["total"], submitted_count=len(status["submitted"]), not_submitted_count=len(status["not_submitted"]), avg_score=avg, practice_link=practice_link, student_url=student_url)


# ==============================================================================
# AI & Generate
# ==============================================================================
@app.route("/ai-slides", methods=["GET", "POST"])
@login_required
def ai_slides():
    # ตรวจสอบ limits (Freemium)
    is_premium = is_premium_user(session["user_id"])
    can_create_topic, topic_msg = UsageLimits.can_create_topic(session["user_id"], is_premium)
    can_ai, ai_msg = UsageLimits.can_ai_generate(session["user_id"], is_premium)

    if request.method == "POST":
        # Check topic limit
        if not can_create_topic:
            flash(f"❌ {topic_msg}", "error")
            return redirect(url_for("pricing"))

        # Check AI limit
        if not can_ai:
            flash(f"❌ {ai_msg}", "error")
            return redirect(url_for("pricing"))

        title = (request.form.get("title") or "").strip()
        if not title:
            flash("Topic title required.", "error")
            return render_template("ai_slides_form.html", can_ai=can_ai, ai_msg=ai_msg)

        bundle = generate_lesson_bundle(
            title=title,
            level=request.form.get("level", "Secondary"),
            language=request.form.get("language", "EN"),
            style=request.form.get("style", "Minimal"),
            text_model="gpt-4o-mini",
        )
        slides = bundle.get("slides", []) or []
        topic = Topic.create(session["user_id"], title, f"AI generated", json.dumps({"slides": slides}, ensure_ascii=False), "ai", None)
        _save_game_and_practice(topic["id"], bundle.get("game") or {}, bundle.get("practice") or [])

        # เพิ่ม AI usage count
        UsageLimits.increment_ai_generate(session["user_id"])

        flash("🎉 สร้างบทเรียนด้วย AI สำเร็จ!", "success")
        return redirect(url_for("topic_detail", topic_id=topic["id"]))

    return render_template("ai_slides_form.html", can_ai=can_ai, ai_msg=ai_msg)


def _extract_text_from_pdf(pdf_path: str) -> str:
    """Extract text from a PDF (pypdf). Raises a friendly error if unreadable."""
    try:
        from pypdf import PdfReader
    except Exception as e:
        raise Exception("Missing dependency: pypdf (ติดตั้งด้วย pip install pypdf)") from e

    reader = PdfReader(pdf_path)
    parts = []
    for page in reader.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception:
            parts.append("")
    text = "\n\n".join(parts).strip()
    if not text:
        raise Exception("PDF has no extractable text (อาจเป็นไฟล์สแกนรูปภาพ)")
    return text


# --- Save helpers for AI generate (slides/game/practice) ---
def _save_game_only(topic_id, game):
    GameQuestion.delete_by_topic(topic_id)
    for set_no in [1, 2, 3]:
        for tile_no, it in enumerate((game.get(str(set_no)) or [])[:24], 1):
            q, a = (it.get("question") or "").strip(), (it.get("answer") or "").strip()
            if q and a: GameQuestion.create(topic_id, set_no, tile_no, q, a, int(it.get("points") or 10))

def _save_practice_only(topic_id, practice):
    PracticeQuestion.delete_by_topic(topic_id)
    for it in (practice or []):
        prompt, choices = (it.get("question") or "").strip(), it.get("choices") or []
        if not prompt or len(choices) != 4: continue
        ci = max(0, min(int(it.get("correct_index") or 0), 3))
        PracticeQuestion.create(topic_id, "multiple_choice", json.dumps({"prompt": prompt, "choices": choices}), str(choices[ci]).strip())

def _save_slides_only(topic_id, slides):
    # Save generated slides to topic.slides_json
    topic = Topic.get_by_id(topic_id)
    if not topic:
        return
    slides_json = json.dumps({"slides": slides or []}, ensure_ascii=False)
    Topic.update(topic_id, topic["name"], topic.get("description") or "", slides_json, topic.get("pdf_file"))

def _save_game_and_practice(topic_id, game, practice):
    _save_game_only(topic_id, game)
    _save_practice_only(topic_id, practice)

def _save_all(topic_id, slides, game, practice):
    # Save slides, game, and practice all at once
    _save_slides_only(topic_id, slides)
    _save_game_only(topic_id, game)
    _save_practice_only(topic_id, practice)


@app.route("/api/topic/<int:topic_id>/generate", methods=["POST"])
@login_required
def api_generate_from_pdf(topic_id):
    # ตรวจสอบ AI limit (Freemium)
    is_premium = is_premium_user(session["user_id"])
    can_ai, ai_msg = UsageLimits.can_ai_generate(session["user_id"], is_premium)

    if not can_ai:
        return jsonify({"ok": False, "error": ai_msg, "upgrade_required": True}), 403

    topic = _get_topic_or_404(topic_id)
    if not topic.get("pdf_file"):
        return _json_error("No PDF", 400)

    mode = ((request.get_json(silent=True) or {}).get("mode") or "all").lower()
    path = os.path.join(app.config["UPLOAD_FOLDER"], topic["pdf_file"])
    if not os.path.exists(path):
        return _json_error("PDF not found", 404)

    try:
        text = _extract_text_from_pdf(path)
    except Exception as e:
        return _json_error(str(e), 400)

    try:
        bundle = generate_lesson_bundle(f"{topic['name']}\n\n[PDF]\n{text[:8000]}", "Secondary", "EN", "Minimal", "gpt-4o-mini")
    except Exception as e:
        return _json_error(str(e), 500)

    # Save based on mode
    if mode == "slides":
        _save_slides_only(topic_id, bundle.get("slides") or [])
    elif mode == "game":
        _save_game_only(topic_id, bundle.get("game") or {})
    elif mode == "practice":
        _save_practice_only(topic_id, bundle.get("practice") or [])
    else:  # mode == "all"
        _save_all(topic_id, bundle.get("slides") or [], bundle.get("game") or {}, bundle.get("practice") or [])

    # เพิ่ม AI usage count
    UsageLimits.increment_ai_generate(session["user_id"])

    # Return stats
    stats = UsageLimits.get_user_stats(session["user_id"])
    return jsonify({
        "ok": True,
        "ai_usage": stats.get("ai_generate_count", 0),
        "ai_limit": UsageLimits.FREE_AI_GENERATE_PER_MONTH if not is_premium else -1
    })

@app.route("/admin")
@admin_required
def admin_dashboard(): return render_template("admin_dashboard.html", topics=Topic.get_all())

@app.route("/admin/topics/create", methods=["GET", "POST"])
@admin_required
def admin_create_topic():
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        if not name: flash("Name required.", "error"); return render_template("admin_create_topic.html")
        topic = Topic.create(session["user_id"], name, request.form.get("description") or "", json.dumps({"slides": []}), "manual", None)
        return redirect(url_for("admin_edit_topic", topic_id=topic["id"]))
    return render_template("admin_create_topic.html")

@app.route("/admin/topics/<int:topic_id>/edit", methods=["GET", "POST"])
@admin_required
def admin_edit_topic(topic_id):
    topic = Topic.get_by_id(topic_id)
    if not topic: return redirect(url_for("admin_dashboard"))
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        slides_json = request.form.get("slides_json") or ""
        try: json.loads(slides_json)
        except: flash("Invalid JSON.", "error"); return render_template("admin_edit_topic.html", topic=topic)
        pdf_filename = topic.get("pdf_file")
        file = request.files.get("pdf_file")
        if file and file.filename and allowed_file(file.filename):
            fn = f"topic{topic_id}_{secrets.token_hex(6)}_{secure_filename(file.filename)}"
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], fn))
            pdf_filename = fn
        Topic.update(topic_id, name, request.form.get("description") or "", slides_json, pdf_filename)
        flash("Saved.", "success")
    return render_template("admin_edit_topic.html", topic=Topic.get_by_id(topic_id))

@app.route("/admin/topics/<int:topic_id>/delete", methods=["POST"])
@admin_required
def admin_delete_topic(topic_id):
    Topic.delete(topic_id)
    return redirect(url_for("admin_dashboard"))

# ==============================================================================
# PATCH: เพิ่ม Short URL Routes สำหรับ QR Code
# เพิ่มใน app.py (ก่อน if __name__ == '__main__')
# ==============================================================================

# Short URLs for QR Code (redirect to public practice pages)
@app.route("/p/<int:topic_id>/mcq")
def qr_practice_mcq(topic_id):
    # Short URL for MCQ practice - redirects to public practice page
    topic = Topic.get_by_id(topic_id)
    if not topic:
        abort(404)
    # Create or get practice link
    link = PracticeLink.get_by_topic(topic_id)
    if not link:
        # Create a new link if doesn't exist
        link = PracticeLink.create(topic_id, topic.get("owner_id") or 1, secrets.token_urlsafe(12), "mcq")
    return redirect(url_for('public_practice', token=link['token']))


@app.route("/p/<int:topic_id>/fill")
def qr_practice_fill(topic_id):
    # Short URL for Fill Blanks practice
    topic = Topic.get_by_id(topic_id)
    if not topic:
        abort(404)
    link = PracticeLink.get_by_topic(topic_id)
    if not link:
        link = PracticeLink.create(topic_id, topic.get("owner_id") or 1, secrets.token_urlsafe(12), "fill")
    return redirect(url_for('public_fill_blanks', token=link['token']))


@app.route("/p/<int:topic_id>/unscramble")
def qr_practice_unscramble(topic_id):
    # Short URL for Unscramble practice
    topic = Topic.get_by_id(topic_id)
    if not topic:
        abort(404)
    link = PracticeLink.get_by_topic(topic_id)
    if not link:
        ink = PracticeLink.create(topic_id, topic.get("owner_id") or 1, secrets.token_urlsafe(12), "unscramble")
    return redirect(url_for('public_unscramble', token=link['token']))

# ==============================================================================
# Library - คลังบทเรียน
# ==============================================================================

@app.route("/library")
@login_required
def library():
    # หน้าแรกคลังบทเรียน - แสดงวิชาทั้งหมด
    subjects = LibrarySubject.get_all_active()
    popular_units = LibraryUnit.get_popular_units(6)
    free_units = LibraryUnit.get_free_units(6)
    is_premium = UserSubscription.is_premium(session["user_id"])
    
    return render_template("library/index.html", 
                           subjects=subjects, 
                           popular_units=popular_units,
                           free_units=free_units,
                           is_premium=is_premium)


@app.route("/library/subject/<int:subject_id>")
@login_required
def library_subject(subject_id):
    # หน้าแสดงบทเรียนในวิชา
    subject = LibrarySubject.get_by_id(subject_id)
    if not subject:
        abort(404)
    
    units = LibraryUnit.get_by_subject(subject_id)
    is_premium = UserSubscription.is_premium(session["user_id"])
    
    # Mark which units user has cloned
    user_clones = {c["unit_id"]: c for c in LibraryClone.get_by_user(session["user_id"])}
    
    return render_template("library/subject.html",
                           subject=subject,
                           units=units,
                           is_premium=is_premium,
                           user_clones=user_clones)


@app.route("/library/unit/<int:unit_id>")
@login_required
def library_unit_detail(unit_id):
    # หน้ารายละเอียดบทเรียน
    unit = LibraryUnit.get_by_id(unit_id)
    if not unit:
        abort(404)
    
    # Increment view count
    LibraryUnit.increment_view(unit_id)
    
    is_premium = UserSubscription.is_premium(session["user_id"])
    can_access = unit["is_free"] == 1 or is_premium
    has_cloned = LibraryClone.has_cloned(session["user_id"], unit_id)
    user_rating = LibraryRating.get_user_rating(session["user_id"], unit_id)
    
    # Parse slides for preview
    slides_preview = []
    if unit.get("slides_json"):
        try:
            slides_data = json.loads(unit["slides_json"])
            slides = slides_data.get("slides", slides_data) if isinstance(slides_data, dict) else slides_data
            # Show limited slides if not premium and not free
            if can_access:
                slides_preview = slides
            else:
                slides_preview = slides[:unit.get("preview_slides", 3)]
        except:
            pass
    
    return render_template("library/unit_detail.html",
                           unit=unit,
                           is_premium=is_premium,
                           can_access=can_access,
                           has_cloned=has_cloned,
                           user_rating=user_rating,
                           slides_preview=slides_preview,
                           total_slides=len(slides_preview) if can_access else "?")


# ==============================================================================
# 2. แก้ไข library_clone_unit (ประมาณบรรทัด 2082-2150)
# ==============================================================================

@app.route("/library/unit/<int:unit_id>/clone", methods=["POST"])
@login_required
def library_clone_unit(unit_id):
    # Clone บทเรียนไปเป็น Topic ของตัวเอง
    unit = LibraryUnit.get_by_id(unit_id)
    if not unit:
        return jsonify({"ok": False, "error": "Unit not found"}), 404
    
    is_premium = UserSubscription.is_premium(session["user_id"])
    can_access = unit["is_free"] == 1 or is_premium
    
    if not can_access:
        return jsonify({"ok": False, "error": "Premium required"}), 403
    
    # Check if already cloned
    if LibraryClone.has_cloned(session["user_id"], unit_id):
        clones = LibraryClone.get_by_user(session["user_id"])
        for c in clones:
            if c["unit_id"] == unit_id:
                return jsonify({"ok": True, "topic_id": c["topic_id"], "already_cloned": True})
    
    # Create new topic from unit
    topic = Topic.create(
        owner_id=session["user_id"],
        name=unit["name"],
        description=unit.get("description") or f"จาก Library: {unit.get('subject_name', '')}",
        slides_json=unit.get("slides_json") or "{}",
        topic_type="library",
        pdf_file=unit.get("pdf_file")
    )
    
    # Copy game questions
    if unit.get("game_json"):
        try:
            game_data = json.loads(unit["game_json"])
            for set_no_str, questions in game_data.items():
                if isinstance(questions, list):
                    for idx, q in enumerate(questions):
                        GameQuestion.create(
                            topic_id=topic["id"],
                            set_no=int(set_no_str),
                            tile_no=q.get("tile_no", idx + 1),
                            question=q.get("question", ""),
                            answer=q.get("answer", ""),
                            points=q.get("points", 10)
                        )
        except Exception as e:
            print(f"Error copying game: {e}")
    
    # Copy practice questions - ใช้ format เดียวกับ _save_practice_only
    if unit.get("practice_json"):
        try:
            practice_data = json.loads(unit["practice_json"])
            if isinstance(practice_data, list):
                for q in practice_data:
                    prompt = q.get("question", "")
                    choices = q.get("choices", [])
                    ci = q.get("correct_index", 0)
                    
                    if prompt and choices:
                        # Format ที่ DB ใช้: question = JSON, correct_answer = string
                        correct_answer = str(choices[ci]).strip() if ci < len(choices) else ""
                        PracticeQuestion.create(
                            topic_id=topic["id"],
                            q_type="multiple_choice",
                            question=json.dumps({"prompt": prompt, "choices": choices}, ensure_ascii=False),
                            correct_answer=correct_answer
                        )
        except Exception as e:
            print(f"Error copying practice: {e}")
    
    # Record clone
    LibraryClone.create(session["user_id"], unit_id, topic["id"])
    
    return jsonify({"ok": True, "topic_id": topic["id"]})


@app.route("/library/unit/<int:unit_id>/rate", methods=["POST"])
@login_required
def library_rate_unit(unit_id):
    # Rate บทเรียน
    unit = LibraryUnit.get_by_id(unit_id)
    if not unit:
        return jsonify({"ok": False, "error": "Unit not found"}), 404
    
    data = request.get_json() or {}
    rating = int(data.get("rating", 0))
    review = (data.get("review") or "").strip()
    
    if rating < 1 or rating > 5:
        return jsonify({"ok": False, "error": "Rating must be 1-5"}), 400
    
    LibraryRating.rate(session["user_id"], unit_id, rating, review)
    
    # Get updated unit
    updated_unit = LibraryUnit.get_by_id(unit_id)
    
    return jsonify({"ok": True, "avg_rating": updated_unit.get("avg_rating", 0)})


@app.route("/library/search")
@login_required
def library_search():
    # ค้นหาบทเรียน
    q = request.args.get("q", "").strip()
    subject_id = request.args.get("subject_id", type=int)
    free_only = request.args.get("free_only") == "1"
    
    results = []
    if q:
        results = LibraryUnit.search(q, subject_id, free_only)
    
    is_premium = UserSubscription.is_premium(session["user_id"])
    subjects = LibrarySubject.get_all_active()
    
    return render_template("library/search.html",
                           query=q,
                           results=results,
                           subjects=subjects,
                           selected_subject=subject_id,
                           free_only=free_only,
                           is_premium=is_premium)


# ==============================================================================
# Premium / Subscription
# ==============================================================================

@app.route("/premium")
@login_required
def premium_page():
    # หน้าแนะนำ Premium
    plans = SubscriptionPlan.get_all_active()
    current_sub = UserSubscription.get_active_subscription(session["user_id"])
    
    return render_template(
        "library/premium.html",
        plans=plans,
        current_sub=current_sub
    )


@app.route("/premium/subscribe/<int:plan_id>", methods=["POST"])
@login_required
def premium_subscribe(plan_id):
    """
    สมัคร Premium (ตัดโหมด Demo ออก)
    - เดิม: สร้าง subscription ทันทีแบบ demo
    - ใหม่: สร้างรายการชำระเงิน แล้วส่ง redirect ไปหน้าชำระเงิน
    """
    user_id = session["user_id"]

    # ✅ อนุญาตให้อัปเกรด/เปลี่ยนแพ็คเกจได้
    # (บล็อกเฉพาะกรณีที่ผู้ใช้มีแพ็คเกจนี้อยู่แล้ว)
    try:
        current_sub = UserSubscription.get_active_subscription(user_id)
    except Exception:
        current_sub = None

    def _sub_plan_id(sub):
        if not sub:
            return None
        if isinstance(sub, dict):
            return sub.get('plan_id') or sub.get('planId') or sub.get('plan')
        return getattr(sub, 'plan_id', None)

    current_plan_id = _sub_plan_id(current_sub)
    if current_plan_id is not None and int(current_plan_id) == int(plan_id):
        return jsonify({"ok": False, "error": "คุณเป็นสมาชิกแพ็คเกจนี้อยู่แล้ว"}), 400

    plan = SubscriptionPlan.get_by_id(plan_id)
    if not plan:
        return jsonify({"ok": False, "error": "ไม่พบแพ็คเกจ"}), 404

    # ถ้ามีรายการค้างชำระอยู่แล้ว ให้เด้งไปต่อ
    pending = PaymentTransaction.get_pending_by_user(user_id, plan_id)
    if pending:
        return jsonify({"ok": True, "redirect": url_for("payment_page", ref_code=pending["reference_code"])})

    # สร้างรายการชำระเงินใหม่
    txn = PaymentTransaction.create(
        user_id=user_id,
        plan_id=plan_id,
        amount=plan["price"]
    )

    return jsonify({"ok": True, "redirect": url_for("payment_page", ref_code=txn["reference_code"])})


# ==============================================================================
# Admin - Library Management
# ==============================================================================

@app.route("/admin/library")
@login_required
@admin_required
def admin_library():
    subjects = LibrarySubject.get_all_active()
    return render_template("admin/library.html", subjects=subjects)


@app.route("/admin/library/subject/create", methods=["GET", "POST"])
@login_required
def admin_library_subject_create():
    # Admin: สร้างวิชาใหม่
    if not _is_admin():
        abort(403)
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("กรุณาใส่ชื่อวิชา", "error")
            return render_template("admin/library_subject_edit.html", subject=None)
        
        subject = LibrarySubject.create(
            name=name,
            description=request.form.get("description", ""),
            grade_level=request.form.get("grade_level", ""),
            subject_type=request.form.get("subject_type", "english"),
            icon=request.form.get("icon", "📚"),
            color=request.form.get("color", "#667eea")
        )
        flash("สร้างวิชาสำเร็จ", "success")
        return redirect(url_for("admin_library"))
    
    return render_template("admin/library_subject_edit.html", subject=None)


@app.route("/admin/library/subject/<int:subject_id>/edit", methods=["GET", "POST"])
@login_required
def admin_library_subject_edit(subject_id):
    # Admin: แก้ไขวิชา
    if not _is_admin():
        abort(403)
    
    subject = LibrarySubject.get_by_id(subject_id)
    if not subject:
        abort(404)
    
    if request.method == "POST":
        LibrarySubject.update(
            subject_id,
            name=request.form.get("name", "").strip(),
            description=request.form.get("description", ""),
            grade_level=request.form.get("grade_level", ""),
            subject_type=request.form.get("subject_type", "english"),
            icon=request.form.get("icon", "📚"),
            color=request.form.get("color", "#667eea")
        )
        flash("บันทึกสำเร็จ", "success")
        return redirect(url_for("admin_library"))
    
    return render_template("admin/library_subject_edit.html", subject=subject)


@app.route("/admin/library/unit/create/<int:subject_id>", methods=["GET", "POST"])
@login_required
def admin_library_unit_create(subject_id):
    # Admin: สร้างบทเรียนใหม่
    if not _is_admin():
        abort(403)
    
    subject = LibrarySubject.get_by_id(subject_id)
    if not subject:
        abort(404)
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if not name:
            flash("กรุณาใส่ชื่อบทเรียน", "error")
            return render_template("admin/library_unit_edit.html", subject=subject, unit=None, slides_count=0, game_count=0, practice_count=0)
        
        # Handle PDF upload
        pdf_filename = None
        if "pdf_file" in request.files:
            pdf = request.files["pdf_file"]
            if pdf and pdf.filename and pdf.filename.endswith(".pdf"):
                pdf_filename = secure_filename(f"lib_{subject_id}_{int(datetime.utcnow().timestamp())}_{pdf.filename}")
                pdf.save(os.path.join(app.config["UPLOAD_FOLDER"], pdf_filename))
        
        unit = LibraryUnit.create(
            subject_id=subject_id,
            name=name,
            unit_number=int(request.form.get("unit_number", 1)),
            description=request.form.get("description", ""),
            is_free=request.form.get("is_free") == "1",
            estimated_time=int(request.form.get("estimated_time", 60)),
            pdf_file=pdf_filename
        )
        flash("สร้างบทเรียนสำเร็จ! ตอนนี้สามารถ Generate เนื้อหาได้", "success")
        return redirect(url_for("admin_library_unit_edit", unit_id=unit["id"]))
    
    return render_template("admin/library_unit_edit.html", subject=subject, unit=None, slides_count=0, game_count=0, practice_count=0)


@app.route("/admin/library/unit/<int:unit_id>/edit", methods=["GET", "POST"])
@login_required
def admin_library_unit_edit(unit_id):
    # Admin: แก้ไขบทเรียน
    if not _is_admin():
        abort(403)
    
    unit = LibraryUnit.get_by_id(unit_id)
    if not unit:
        abort(404)
    
    subject = LibrarySubject.get_by_id(unit["subject_id"])
    
    if request.method == "POST":
        # Handle PDF upload
        pdf_filename = unit.get("pdf_file")
        if "pdf_file" in request.files:
            pdf = request.files["pdf_file"]
            if pdf and pdf.filename and pdf.filename.endswith(".pdf"):
                pdf_filename = secure_filename(f"lib_{unit['subject_id']}_{int(datetime.utcnow().timestamp())}_{pdf.filename}")
                pdf.save(os.path.join(app.config["UPLOAD_FOLDER"], pdf_filename))
        
        # Get JSON from form (if edited manually)
        slides_json = request.form.get("slides_json", unit.get("slides_json") or "{}")
        game_json = request.form.get("game_json", unit.get("game_json") or "{}")
        practice_json = request.form.get("practice_json", unit.get("practice_json") or "[]")
        
        LibraryUnit.update(
            unit_id,
            name=request.form.get("name", "").strip(),
            unit_number=int(request.form.get("unit_number", 1)),
            description=request.form.get("description", ""),
            is_free=1 if request.form.get("is_free") == "1" else 0,
            estimated_time=int(request.form.get("estimated_time", 60)),
            pdf_file=pdf_filename,
            slides_json=slides_json,
            game_json=game_json,
            practice_json=practice_json
        )
        flash("บันทึกสำเร็จ", "success")
        return redirect(url_for("admin_library_unit_edit", unit_id=unit_id))
    
    # Reload unit after potential update
    unit = LibraryUnit.get_by_id(unit_id)
    
    # Calculate content stats
    slides_count = 0
    game_count = 0
    practice_count = 0
    
    if unit.get("slides_json"):
        try:
            slides_data = json.loads(unit["slides_json"])
            if isinstance(slides_data, dict):
                slides = slides_data.get("slides", [])
            else:
                slides = slides_data
            slides_count = len(slides) if isinstance(slides, list) else 0
        except:
            pass
    
    if unit.get("game_json"):
        try:
            game_data = json.loads(unit["game_json"])
            for questions in game_data.values():
                if isinstance(questions, list):
                    game_count += len(questions)
        except:
            pass
    
    if unit.get("practice_json"):
        try:
            practice_data = json.loads(unit["practice_json"])
            practice_count = len(practice_data) if isinstance(practice_data, list) else 0
        except:
            pass
    
    return render_template("admin/library_unit_edit.html", 
                           subject=subject, 
                           unit=unit,
                           slides_count=slides_count,
                           game_count=game_count,
                           practice_count=practice_count)

@app.route("/admin/library/unit/<int:unit_id>/generate/<gen_type>", methods=["POST"])
@login_required
def admin_library_unit_generate(unit_id, gen_type):
    # Admin: Generate เนื้อหาด้วย AI
    if not _is_admin():
        return jsonify({"ok": False, "error": "Forbidden"}), 403
    
    unit = LibraryUnit.get_by_id(unit_id)
    if not unit:
        return jsonify({"ok": False, "error": "Unit not found"}), 404
    
    if gen_type not in ["all", "slides", "game", "practice"]:
        return jsonify({"ok": False, "error": "Invalid type"}), 400
    
    topic_name = unit["name"]
    
    try:
        # Generate content using AI
        bundle = generate_lesson_bundle(topic_name)
        
        if not bundle:
            return jsonify({"ok": False, "error": "AI generation failed"}), 500
        
        # Prepare updates
        updates = {}
        result = {"ok": True}
        
        # Slides
        if gen_type in ["all", "slides"]:
            slides = bundle.get("slides") or []
            slides_json = json.dumps({"slides": slides}, ensure_ascii=False)
            updates["slides_json"] = slides_json
            result["slides_count"] = len(slides)
            result["slides_json"] = slides_json
        
        # Game - convert from AI format to library format
        if gen_type in ["all", "game"]:
            game_raw = bundle.get("game") or {}
            game_data = {}
            total_game = 0
            
            for set_no in [1, 2, 3]:
                set_key = str(set_no)
                questions = game_raw.get(set_key) or game_raw.get(f"set{set_no}") or []
                if questions:
                    game_data[set_key] = []
                    for idx, q in enumerate(questions):
                        game_data[set_key].append({
                            "tile_no": idx + 1,
                            "question": q.get("question", ""),
                            "answer": q.get("answer", ""),
                            "points": q.get("points", 10)
                        })
                        total_game += 1
            
            game_json = json.dumps(game_data, ensure_ascii=False)
            updates["game_json"] = game_json
            result["game_count"] = total_game
            result["game_json"] = game_json
        
        # Practice - convert from AI format to library format
        if gen_type in ["all", "practice"]:
            practice_raw = bundle.get("practice") or []
            practice_data = []
            
            for q in practice_raw:
                practice_data.append({
                    "question": q.get("question", ""),
                    "choices": q.get("choices", []),
                    "correct_index": q.get("correct_index", 0)
                })
            
            practice_json = json.dumps(practice_data, ensure_ascii=False)
            updates["practice_json"] = practice_json
            result["practice_count"] = len(practice_data)
            result["practice_json"] = practice_json
        
        # Save to database
        if updates:
            LibraryUnit.update(unit_id, **updates)
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ==============================================================================
# 1. แก้ไข admin_library_import_from_topic (ประมาณบรรทัด 2369-2403)
# ==============================================================================

@app.route("/admin/library/unit/<int:unit_id>/import-from-topic/<int:topic_id>", methods=["POST"])
@login_required
def admin_library_import_from_topic(unit_id, topic_id):
    # Admin: Import เนื้อหาจาก Topic ที่มีอยู่
    if not _is_admin():
        abort(403)
    
    unit = LibraryUnit.get_by_id(unit_id)
    topic = Topic.get_by_id(topic_id)
    
    if not unit or not topic:
        return jsonify({"ok": False, "error": "Not found"}), 404
    
    # Copy slides
    slides_json = topic.get("slides_json", "{}")
    
    # Copy game questions
    game_data = {}
    for set_no in [1, 2, 3]:
        questions = GameQuestion.get_by_topic_and_set(topic_id, set_no)
        if questions:
            game_data[str(set_no)] = [
                {
                    "tile_no": q.get("tile_no", idx + 1),
                    "question": q["question"], 
                    "answer": q["answer"], 
                    "points": q.get("points", 10)
                } 
                for idx, q in enumerate(questions)
            ]
    
    # Copy practice questions
    # Format ใน DB: question = JSON{"prompt": "...", "choices": [...]}, correct_answer = "คำตอบที่ถูก"
    practice_questions = PracticeQuestion.get_by_topic(topic_id)
    practice_data = []
    for q in practice_questions:
        try:
            # Parse question JSON
            q_data = json.loads(q["question"])
            prompt = q_data.get("prompt", "")
            choices = q_data.get("choices", [])
            correct_answer = q.get("correct_answer", "")
            
            # Find correct_index
            correct_index = 0
            for idx, choice in enumerate(choices):
                if str(choice).strip() == str(correct_answer).strip():
                    correct_index = idx
                    break
            
            practice_data.append({
                "question": prompt,
                "choices": choices,
                "correct_index": correct_index
            })
        except:
            # Fallback for old format
            practice_data.append({
                "question": q["question"],
                "choices": [],
                "correct_index": 0
            })
    
    LibraryUnit.update(
        unit_id,
        slides_json=slides_json,
        game_json=json.dumps(game_data, ensure_ascii=False) if game_data else "",
        practice_json=json.dumps(practice_data, ensure_ascii=False) if practice_data else ""
    )
    
    return jsonify({"ok": True})


# ==============================================================================
# Pricing & Limits (Freemium)
# ==============================================================================
@app.route("/pricing")
def pricing():
    # หน้าแสดงราคาและแพ็คเกจ
    plans = SubscriptionPlan.get_all_active()

    user_id = session.get("user_id")
    is_premium = False
    current_sub = None
    stats = {"topic_count": 0, "classroom_count": 0, "ai_generate_count": 0}

    if user_id:
        is_premium = is_premium_user(user_id)
        current_sub = UserSubscription.get_active_subscription(user_id)
        try:
            stats = UsageLimits.get_user_stats(user_id)
        except Exception:
            pass

    return render_template(
        "pricing.html",
        plans=plans,
        is_premium=is_premium,
        current_sub=current_sub,
        stats=stats,
        free_limits={
            "topics": getattr(UsageLimits, "FREE_TOPICS", 5),
            "classrooms": getattr(UsageLimits, "FREE_CLASSROOMS", 2),
            "ai_generate": getattr(UsageLimits, "FREE_AI_GENERATE_PER_MONTH", 3),
        },
    )


@app.route("/api/user/limits")
@login_required
def api_user_limits():
    # API: ดึงข้อมูล limits ของ user
    user_id = session["user_id"]
    is_premium = is_premium_user(user_id)
    stats = UsageLimits.get_user_stats(user_id)

    return jsonify({
        "ok": True,
        "is_premium": is_premium,
        "stats": stats,
        "limits": {
            "topics": -1 if is_premium else UsageLimits.FREE_TOPICS,
            "classrooms": -1 if is_premium else UsageLimits.FREE_CLASSROOMS,
            "ai_generate": -1 if is_premium else UsageLimits.FREE_AI_GENERATE_PER_MONTH,
        },
        "can_create_topic": UsageLimits.can_create_topic(user_id, is_premium)[0],
        "can_create_classroom": UsageLimits.can_create_classroom(user_id, is_premium)[0],
        "can_ai_generate": UsageLimits.can_ai_generate(user_id, is_premium)[0],
    })

# ==============================================================================
# Errors
# ==============================================================================
@app.errorhandler(403)
def forbidden(e): return (jsonify({"ok": False, "error": "Forbidden"}), 403) if _wants_json_response() else (render_template("error.html", error_code=403, error_msg="ไม่มีสิทธิ์"), 403)
@app.errorhandler(404)
def not_found(e): return (jsonify({"ok": False, "error": "Not found"}), 404) if _wants_json_response() else (render_template("error.html", error_code=404, error_msg="ไม่พบหน้านี้"), 404)
@app.errorhandler(500)
def server_error(e): return (jsonify({"ok": False, "error": "Server error"}), 500) if _wants_json_response() else (render_template("error.html", error_code=500, error_msg="เกิดข้อผิดพลาด"), 500)


def _crc16_ccitt(data: bytes) -> int:
    """Calculate CRC16-CCITT (XModem)"""
    crc = 0xFFFF
    for byte in data:
        crc ^= byte << 8
        for _ in range(8):
            if crc & 0x8000:
                crc = (crc << 1) ^ 0x1021
            else:
                crc <<= 1
            crc &= 0xFFFF
    return crc


def generate_promptpay_qr_payload(promptpay_id: str, amount: float) -> str:
    """
    Generate PromptPay EMVCo QR payload
    promptpay_id: เลขบัตรประชาชน 13 หลัก หรือ เบอร์โทร 10 หลัก
    amount: จำนวนเงิน (บาท)
    """
    promptpay_id = promptpay_id.replace("-", "").replace(" ", "")
    
    def tlv(tag: str, value: str) -> str:
        return f"{tag}{len(value):02d}{value}"
    
    # Determine ID type and format
    if len(promptpay_id) == 10:  # Phone number
        # Format: 0066 + 9 digits (remove leading 0)
        formatted_id = "0066" + promptpay_id[1:]
        id_tag = "01"  # Phone
    elif len(promptpay_id) == 13:  # National ID
        formatted_id = promptpay_id
        id_tag = "02"  # National ID
    else:
        raise ValueError("Invalid PromptPay ID")
    
    # Build Merchant Account Information (Tag 29)
    aid = tlv("00", "A000000677010111")  # PromptPay AID
    mobile_or_id = tlv(id_tag, formatted_id)
    merchant_info = tlv("29", aid + mobile_or_id)
    
    # Build full payload
    payload = ""
    payload += tlv("00", "01")           # Payload Format Indicator
    payload += tlv("01", "12")           # Point of Initiation Method (12 = Dynamic)
    payload += merchant_info             # Merchant Account (Tag 29)
    payload += tlv("52", "0000")         # Merchant Category Code
    payload += tlv("53", "764")          # Transaction Currency (764 = THB)
    payload += tlv("54", f"{amount:.2f}") # Transaction Amount
    payload += tlv("58", "TH")           # Country Code
    payload += "6304"                    # CRC placeholder
    
    # Calculate and append CRC
    crc = _crc16_ccitt(payload.encode('utf-8'))
    payload = payload[:-4] + f"6304{crc:04X}"
    
    return payload


def get_promptpay_qr_image_url(promptpay_id: str, amount: float, size: int = 300) -> str:
    """Generate QR code image URL using free API"""
    payload = generate_promptpay_qr_payload(promptpay_id, amount)
    encoded = urllib.parse.quote(payload)
    return f"https://api.qrserver.com/v1/create-qr-code/?size={size}x{size}&data={encoded}"

    
def verify_slip_with_easyslip(image_base64: str, api_key: str) -> dict:
    """
    ส่งสลิปไปตรวจสอบที่ EasySlip API (Verify By Base64)
    - ส่ง checkDuplicate=True เพื่อให้ EasySlip ช่วยกันสลิปซ้ำระดับ API
    Returns:
      {
        "success": bool,
        "data": dict|None,      # payload ที่ EasySlip ส่งกลับ (ถ้ามี)
        "error": str|None,      # message เช่น duplicate_slip / invalid_image / unauthorized
        "status": int|None      # status code จาก EasySlip (ถ้ามี)
      }
    """
    try:
        response = requests.post(
            "https://developer.easyslip.com/api/v1/verify",
            json={"image": image_base64, "checkDuplicate": True},
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            timeout=30,
        )

        # EasySlip จะส่งรูปแบบ {status:200, data:{...}} หรือ {status:400, message:'...'}
        try:
            result = response.json()
        except Exception:
            result = {}

        if response.status_code == 200 and result.get("status") == 200:
            return {
                "success": True,
                "data": result.get("data", {}) or {},
                "error": None,
                "status": result.get("status", 200),
            }

        # กรณี error บางแบบ (เช่น duplicate_slip) ยังมี data กลับมา
        return {
            "success": False,
            "data": result.get("data"),
            "error": result.get("message", f"HTTP {response.status_code}"),
            "status": result.get("status", response.status_code),
        }

    except requests.Timeout:
        return {"success": False, "data": None, "error": "Request timeout", "status": None}
    except Exception as e:
        return {"success": False, "data": None, "error": str(e), "status": None}


def validate_slip_amount(slip_data: dict, expected_amount: float, tolerance: float = 0.0) -> dict:
    """
    ตรวจสอบจำนวนเงินในสลิป
    tolerance: ความคลาดเคลื่อนที่ยอมรับได้ (บาท)
    """
    try:
        slip_amount = float(slip_data.get("amount", {}).get("amount", 0))
    except Exception:
        slip_amount = 0.0

    is_valid = slip_amount >= (expected_amount - tolerance)

    return {
        "valid": is_valid,
        "slip_amount": slip_amount,
        "expected_amount": expected_amount,
        "difference": slip_amount - expected_amount,
        "error": None if is_valid else f"จำนวนเงินไม่ตรง (โอน {slip_amount:.2f} / ต้องการ {expected_amount:.2f} บาท)"
    }
    
@app.route("/payment/create/<int:plan_id>", methods=["POST"])
@login_required
def payment_create(plan_id):
    """สร้างรายการชำระเงินใหม่"""
    user_id = session["user_id"]

    # ✅ อนุญาตให้อัปเกรด/เปลี่ยนแพ็คเกจได้
    # (บล็อกเฉพาะกรณีที่ผู้ใช้มีแพ็คเกจนี้อยู่แล้ว)
    try:
        current_sub = UserSubscription.get_active_subscription(user_id)
    except Exception:
        current_sub = None

    def _sub_plan_id(sub):
        if not sub:
            return None
        if isinstance(sub, dict):
            return sub.get('plan_id') or sub.get('planId') or sub.get('plan')
        return getattr(sub, 'plan_id', None)

    current_plan_id = _sub_plan_id(current_sub)
    if current_plan_id is not None and int(current_plan_id) == int(plan_id):
        return jsonify({"ok": False, "error": "คุณเป็นสมาชิกแพ็คเกจนี้อยู่แล้ว"}), 400
    
    # หา plan
    plan = SubscriptionPlan.get_by_id(plan_id)
    if not plan:
        return jsonify({"ok": False, "error": "ไม่พบแพ็คเกจ"}), 404
    
    # ตรวจสอบว่ามี pending transaction อยู่หรือไม่
    pending = PaymentTransaction.get_pending_by_user(user_id, plan_id)
    if pending:
        return jsonify({
            "ok": True,
            "redirect": url_for("payment_page", ref_code=pending["reference_code"])
        })
    
    # สร้าง transaction ใหม่
    txn = PaymentTransaction.create(
        user_id=user_id,
        plan_id=plan_id,
        amount=plan["price"]
    )
    
    return jsonify({
        "ok": True,
        "redirect": url_for("payment_page", ref_code=txn["reference_code"])
    })


# ------------------------------------------------------------------------------
# Payment slip validation helpers + lightweight DB migration (SQLite)
# ------------------------------------------------------------------------------

def _normalize_name(s: str) -> str:
    s = (s or '').strip().lower()
    # remove spaces and common punctuation
    return re.sub(r'[\s\-\._,:\'\"\(\)\[\]\{\}]', '', s)


def _extract_slip_trans_ref(slip_data: dict) -> str:
    """พยายามดึงรหัสธุรกรรมที่ไม่ซ้ำ (transRef/transactionRef) จาก EasySlip payload"""
    if not isinstance(slip_data, dict):
        return ''
    # รองรับหลายรูปแบบ
    for k in ("transRef", "trans_ref", "transactionRef", "transaction_ref", "reference", "ref"):
        v = slip_data.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    # บาง payload ซ่อนอยู่ใน data/transaction
    for path in (
        ("data", "transRef"),
        ("data", "transactionRef"),
        ("transaction", "transRef"),
        ("transaction", "transactionRef"),
    ):
        cur = slip_data
        ok = True
        for p in path:
            if isinstance(cur, dict) and p in cur:
                cur = cur[p]
            else:
                ok = False
                break
        if ok and isinstance(cur, str) and cur.strip():
            return cur.strip()
    return ''


def _extract_receiver_info(slip_data: dict) -> tuple[str, str]:
    """คืนค่า (receiver_name, receiver_id) จากผล EasySlip (รองรับโครงสร้าง nested)

    จากเอกสาร EasySlip (verify by base64) จะอยู่ประมาณนี้:
      receiver.account.name.th / receiver.account.name.en
      receiver.account.proxy.account หรือ receiver.account.bank.account (มักถูก mask ด้วย x)
    """
    if not isinstance(slip_data, dict):
        return ("", "")

    # ช่วยดึงค่า nested แบบปลอดภัย
    def _get(obj, *path, default=None):
        cur = obj
        for p in path:
            if isinstance(cur, dict) and p in cur:
                cur = cur[p]
            else:
                return default
        return cur

    # receiver อาจอยู่ชั้นบน หรือใน data
    receiver = None
    if isinstance(slip_data.get("receiver"), dict):
        receiver = slip_data.get("receiver")
    elif isinstance(slip_data.get("data"), dict) and isinstance(slip_data["data"].get("receiver"), dict):
        receiver = slip_data["data"].get("receiver")

    # fallback แบบเก่า
    if receiver is None:
        for k in ("to", "creditor"):
            if isinstance(slip_data.get(k), dict):
                receiver = slip_data.get(k); break
        if receiver is None and isinstance(slip_data.get("data"), dict):
            for k in ("to", "creditor"):
                if isinstance(slip_data["data"].get(k), dict):
                    receiver = slip_data["data"].get(k); break

    recv_name = ""
    recv_id = ""

    if isinstance(receiver, dict):
        # name: receiver.account.name.th/en
        th = _get(receiver, "account", "name", "th")
        en = _get(receiver, "account", "name", "en")
        for v in (th, en):
            if isinstance(v, str) and v.strip():
                recv_name = v.strip()
                break

        # id: proxy.account หรือ bank.account (มัก mask)
        proxy_acc = _get(receiver, "account", "proxy", "account")
        bank_acc = _get(receiver, "account", "bank", "account")
        for v in (proxy_acc, bank_acc):
            if isinstance(v, str) and v.strip():
                recv_id = v.strip()
                break

        # รองรับ key รูปแบบอื่น ๆ (กันกรณี response เปลี่ยน)
        if not recv_name:
            for nk in ("name", "accountName", "displayName"):
                v = receiver.get(nk)
                if isinstance(v, str) and v.strip():
                    recv_name = v.strip(); break
        if not recv_id:
            for ik in ("id", "account", "accountNo", "promptpay", "promptpayId"):
                v = receiver.get(ik)
                if isinstance(v, str) and v.strip():
                    recv_id = v.strip(); break

    # fallback: บาง payload ใส่ name / id ไว้ระดับบน
    if not recv_name:
        for nk in ("receiverName", "toName"):
            v = slip_data.get(nk)
            if isinstance(v, str) and v.strip():
                recv_name = v.strip(); break
        if not recv_name and isinstance(slip_data.get("data"), dict):
            for nk in ("receiverName", "toName"):
                v = slip_data["data"].get(nk)
                if isinstance(v, str) and v.strip():
                    recv_name = v.strip(); break

    if not recv_id:
        for ik in ("receiverId", "toId", "promptpay"):
            v = slip_data.get(ik)
            if isinstance(v, str) and v.strip():
                recv_id = v.strip(); break
        if not recv_id and isinstance(slip_data.get("data"), dict):
            for ik in ("receiverId", "toId", "promptpay"):
                v = slip_data["data"].get(ik)
                if isinstance(v, str) and v.strip():
                    recv_id = v.strip(); break

    return (recv_name, recv_id)


def _masked_id_match(expected_promptpay: str, got_masked: str) -> bool:
    """เทียบ expected (เลขเต็ม) กับ got ที่ถูก mask ด้วย x/* เช่น 123xxxxxxxx4567"""
    exp = re.sub(r"\D", "", expected_promptpay or "")
    got = (got_masked or "").strip()
    if not exp or not got:
        return False

    # เอาเฉพาะเลข + ตัว mask
    got_clean = re.sub(r"[^0-9xX\*]", "", got)
    got_clean = got_clean.replace("*", "x").replace("X", "x")

    if "x" not in got_clean:
        # ไม่มี mask → เทียบตรง
        return re.sub(r"\D", "", got_clean) == exp

    # แยก prefix/suffix ตัวเลข
    prefix = re.match(r"^\d+", got_clean)
    suffix = re.search(r"\d+$", got_clean)
    pre = prefix.group(0) if prefix else ""
    suf = suffix.group(0) if suffix else ""

    if pre and not exp.startswith(pre):
        return False
    if suf and not exp.endswith(suf):
        return False
    return True


def validate_slip_receiver(slip_data: dict, expected_name: str, expected_promptpay: str) -> dict:
    """เช็คชื่อผู้รับ + promptpay id ให้ตรง (รองรับกรณีเลขถูก mask ในสลิป)"""
    recv_name, recv_id = _extract_receiver_info(slip_data)

    # ชื่อผู้รับ: รองรับหลายชื่อ (aliases) แยกด้วย |
    aliases = [x.strip() for x in (os.environ.get("PROMPTPAY_NAME_ALIASES", "")).split("|") if x.strip()]
    candidates = [expected_name] + aliases if expected_name else aliases

    rn = _normalize_name(recv_name)
    name_ok = False
    for cand in candidates:
        en = _normalize_name(cand)
        if en and rn and (en in rn or rn in en):
            name_ok = True
            break

    if not candidates:
        name_ok = True  # ถ้าไม่ได้ตั้งชื่อไว้ ก็ไม่บังคับ

    # id: ถ้าในสลิปเป็น mask ให้เทียบแบบ prefix/suffix
    exp_digits = re.sub(r"\D", "", expected_promptpay or "")
    got_raw = (recv_id or "").strip()

    id_ok = True
    if exp_digits:
        if not got_raw:
            id_ok = False
        else:
            # ถ้าเห็น mask ก็ใช้ masked match
            if any(ch in got_raw for ch in ("x", "X", "*")):
                id_ok = _masked_id_match(exp_digits, got_raw)
            else:
                id_ok = re.sub(r"\D", "", got_raw) == exp_digits

    if not name_ok:
        return {"valid": False, "error": f"ชื่อผู้รับไม่ตรง (ในสลิป: {recv_name or 'ไม่พบชื่อผู้รับ'})"}
    if not id_ok:
        return {"valid": False, "error": f"PromptPay/บัญชีผู้รับไม่ตรง (ในสลิป: {recv_id or 'ไม่พบเลขผู้รับ'})"}

    return {"valid": True, "error": None}


def _ensure_payment_schema() -> None:
    """เพิ่มคอลัมน์ที่จำเป็นสำหรับกันสลิปซ้ำ/เช็คผู้รับ (ปลอดภัย: ไม่ลบข้อมูลเดิม)"""
    try:
        conn = get_db()
        c = conn.cursor()

        # ตารางจริงของคุณคือ payment_transactions (ไม่ใช่ payment_transactionss)
        c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='payment_transactions'")
        if not c.fetchone():
            conn.close()
            return

        c.execute("PRAGMA table_info(payment_transactions)")
        cols = {row[1] for row in c.fetchall()}

        # เพิ่มคอลัมน์ถ้ายังไม่มี
        if "slip_trans_ref" not in cols:
            c.execute("ALTER TABLE payment_transactions ADD COLUMN slip_trans_ref TEXT")
        if "receiver_name" not in cols:
            c.execute("ALTER TABLE payment_transactions ADD COLUMN receiver_name TEXT")
        if "receiver_id" not in cols:
            c.execute("ALTER TABLE payment_transactions ADD COLUMN receiver_id TEXT")

        # แนะนำ: ทำ unique index กันสลิปซ้ำระดับ DB ด้วย (ปลอดภัย)
        try:
            c.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS uq_payment_slip_trans_ref
                ON payment_transactions(slip_trans_ref)
                WHERE slip_trans_ref IS NOT NULL AND slip_trans_ref <> ''
            """)
        except Exception:
            pass

        conn.commit()
        conn.close()
    except Exception:
        # ไม่ให้พังตอนรัน ถ้า migration ล้มเหลว
        pass


def _slip_trans_ref_used(trans_ref: str, exclude_txn_id: Optional[int] = None) -> bool:
    if not trans_ref:
        return False

    _ensure_payment_schema()

    conn = get_db()
    c = conn.cursor()

    if exclude_txn_id is None:
        c.execute(
            "SELECT id FROM payment_transactions WHERE slip_trans_ref = ? LIMIT 1",
            (trans_ref,)
        )
    else:
        c.execute(
            "SELECT id FROM payment_transactions WHERE slip_trans_ref = ? AND id <> ? LIMIT 1",
            (trans_ref, exclude_txn_id)
        )

    row = c.fetchone()
    conn.close()
    return row is not None


@app.route("/payment/<ref_code>")
@login_required
def payment_page(ref_code):
    """หน้าชำระเงิน - แสดง QR Code"""
    txn = PaymentTransaction.get_by_reference(ref_code)
    
    if not txn:
        flash("ไม่พบรายการชำระเงิน", "error")
        return redirect(url_for("pricing"))
    
    if txn["user_id"] != session["user_id"]:
        abort(403)
    
    if txn["status"] == "completed":
        flash("ชำระเงินเรียบร้อยแล้ว!", "success")
        return redirect(url_for("dashboard"))
    
    # สร้าง QR Code URL
    qr_url = get_promptpay_qr_image_url(PROMPTPAY_ID, txn["amount"])
    
    return render_template("payment.html",
                           txn=txn,
                           qr_url=qr_url,
                           promptpay_id=PROMPTPAY_ID,
                           promptpay_name=PROMPTPAY_NAME)


@app.route("/payment/<ref_code>/verify", methods=["POST"])
@login_required
def payment_verify(ref_code):
    """อัปโหลดและตรวจสอบสลิป"""
    txn = PaymentTransaction.get_by_reference(ref_code)
    
    if not txn:
        return jsonify({"ok": False, "error": "ไม่พบรายการ"}), 404
    
    if txn["user_id"] != session["user_id"]:
        return jsonify({"ok": False, "error": "ไม่มีสิทธิ์"}), 403
    
    if txn["status"] == "completed":
        return jsonify({"ok": False, "error": "ชำระเงินเรียบร้อยแล้ว"}), 400
    
    # รับไฟล์สลิป
    if "slip" not in request.files:
        return jsonify({"ok": False, "error": "กรุณาอัปโหลดสลิป"}), 400
    
    slip_file = request.files["slip"]
    if not slip_file or not slip_file.filename:
        return jsonify({"ok": False, "error": "กรุณาเลือกไฟล์"}), 400
    
    # อ่านและ encode เป็น base64
    slip_bytes = slip_file.read()
    slip_base64 = base64.b64encode(slip_bytes).decode("utf-8")
    
    # บันทึกไฟล์สลิป
    slip_filename = f"slip_{ref_code}_{secrets.token_hex(4)}.jpg"
    slip_path = os.path.join(app.config["UPLOAD_FOLDER"], slip_filename)
    with open(slip_path, "wb") as f:
        f.write(slip_bytes)
    
    # อัปเดตสถานะเป็น verifying
    PaymentTransaction.update_status(txn["id"], "verifying", slip_image=slip_filename)
    
    # ส่งไปตรวจสอบที่ EasySlip
    easyslip_result = verify_slip_with_easyslip(slip_base64, EASYSLIP_API_KEY)
    
    
    if not easyslip_result.get("success"):
        # บางกรณี (เช่น duplicate_slip) ยังมี data กลับมา → เก็บ transRef ไว้กันสลิปซ้ำ
        slip_data = easyslip_result.get("data") or {}
        _ensure_payment_schema()
        trans_ref = _extract_slip_trans_ref(slip_data) if isinstance(slip_data, dict) else ""

        # เก็บ error ลง DB (และเก็บ transRef ถ้าหาได้)
        try:
            err_payload = {"error": easyslip_result.get("error")}
            if trans_ref:
                err_payload["transRef"] = trans_ref
            PaymentTransaction.update_status(txn["id"], "failed", easyslip_data=json.dumps(err_payload))
            if trans_ref:
                # อัปเดต slip_trans_ref ไว้ด้วย (กันคนพยายามส่งซ้ำ)
                conn = get_db()
                c = conn.cursor()
                c.execute("UPDATE payment_transactions SET slip_trans_ref = COALESCE(?, slip_trans_ref) WHERE id = ?", (trans_ref, txn["id"]))
                conn.commit()
                conn.close()
        except Exception:
            pass

        # ข้อความที่เป็นมิตรกับผู้ใช้
        if (easyslip_result.get("error") or "").strip() == "duplicate_slip":
            return jsonify({"ok": False, "error": "สลิปนี้ถูกใช้ไปแล้ว (duplicate slip)"}), 400

        return jsonify({
            "ok": False,
            "error": f"ไม่สามารถตรวจสอบสลิปได้: {easyslip_result.get('error')}"
        }), 400
    # ตรวจสอบจำนวนเงิน

    slip_data = easyslip_result["data"]

    # ✅ กันสลิปซ้ำ: ใช้ transRef จาก EasySlip
    _ensure_payment_schema()
    trans_ref = _extract_slip_trans_ref(slip_data)
    if trans_ref and _slip_trans_ref_used(trans_ref, exclude_txn_id=txn["id"]):
        PaymentTransaction.update_status(txn["id"], "failed", easyslip_data=json.dumps({"error": "duplicate_slip", "transRef": trans_ref}))
        return jsonify({"ok": False, "error": "สลิปนี้ถูกใช้แล้ว (ห้ามใช้สลิปซ้ำ)"}), 400

    # ✅ เช็คชื่อผู้รับ + PromptPay ID
    recv_check = validate_slip_receiver(slip_data, PROMPTPAY_NAME, PROMPTPAY_ID)
    if not recv_check["valid"]:
        PaymentTransaction.update_status(txn["id"], "failed", easyslip_data=json.dumps({"error": recv_check["error"], "transRef": trans_ref}))
        return jsonify({"ok": False, "error": recv_check["error"]}), 400

    validation = validate_slip_amount(slip_data, txn["amount"])
    
    # บันทึกข้อมูล EasySlip
    PaymentTransaction.update_status(txn["id"], 
                                     "completed" if validation["valid"] else "failed",
                                     easyslip_data=json.dumps(slip_data))
    
    if not validation["valid"]:
        return jsonify({"ok": False, "error": validation["error"]}), 400
    
    # ✅ สำเร็จ - สร้าง Subscription
    plan = SubscriptionPlan.get_by_id(txn["plan_id"])
    UserSubscription.create(
        user_id=txn["user_id"],
        plan_id=txn["plan_id"],
        duration_days=plan["duration_days"],
        payment_ref=ref_code
    )
    
    return jsonify({
        "ok": True,
        "message": "🎉 ชำระเงินสำเร็จ! คุณเป็น Premium แล้ว",
        "redirect": url_for("dashboard")
    })


@app.route("/admin/payments")
@login_required
def admin_payments():
    """Admin: ดูรายการชำระเงินทั้งหมด"""
    if not _is_admin():
        abort(403)
    
    transactions = PaymentTransaction.get_all_for_admin(100)
    return render_template("admin/payments.html", transactions=transactions)


@app.route("/admin/users")
@login_required
def admin_users():
    """Admin: ดูรายชื่อผู้ใช้ + จำนวนผู้ใช้ทั้งหมด + สถานะ Premium/วันหมดอายุ"""
    if not _is_admin():
        abort(403)

    q = (request.args.get("q") or "").strip()
    now = datetime.utcnow().isoformat()

    conn = get_db()
    c = conn.cursor()

    # จำนวนผู้ใช้ทั้งหมด
    c.execute("SELECT COUNT(*) FROM users")
    row = c.fetchone()
    total_users = int(row[0]) if row and row[0] is not None else 0

    # จำนวน Premium ที่ยังไม่หมดอายุ
    try:
        c.execute("""
            SELECT COUNT(DISTINCT user_id)
            FROM user_subscriptions
            WHERE status='active' AND expires_at > ?
        """, (now,))
        pr = c.fetchone()
        total_premium_active = int(pr[0]) if pr and pr[0] is not None else 0
    except Exception:
        total_premium_active = 0

    # รายการผู้ใช้ + subscription ล่าสุด (ถ้ามี)
    base_sql = """
        SELECT
            u.id, u.email, u.role, u.created_at,
            us.id as sub_id, us.plan_id, us.status as sub_status, us.started_at, us.expires_at, us.payment_ref,
            sp.name as plan_name, sp.price as plan_price
        FROM users u
        LEFT JOIN (
            SELECT s1.*
            FROM user_subscriptions s1
            JOIN (
                SELECT user_id, MAX(expires_at) AS max_exp
                FROM user_subscriptions
                GROUP BY user_id
            ) mx
            ON s1.user_id = mx.user_id AND s1.expires_at = mx.max_exp
        ) us ON us.user_id = u.id
        LEFT JOIN subscription_plans sp ON sp.id = us.plan_id
    """

    if q:
        c.execute(base_sql + " WHERE u.email LIKE ? ORDER BY u.id DESC LIMIT 300", (f"%{q}%",))
    else:
        c.execute(base_sql + " ORDER BY u.id DESC LIMIT 300")

    rows = c.fetchall()
    conn.close()

    users = []
    for r in rows:
        expires_at = r["expires_at"] if "expires_at" in r.keys() else None
        is_premium = False
        if expires_at:
            try:
                is_premium = (r["sub_status"] == "active") and (expires_at > now)
            except Exception:
                is_premium = False

        users.append({
            "id": r["id"],
            "email": r["email"],
            "role": r["role"] or "",
            "created_at": r["created_at"] or "",
            "sub_id": r["sub_id"],
            "plan_id": r["plan_id"],
            "plan_name": r["plan_name"] or "",
            "plan_price": r["plan_price"] if "plan_price" in r.keys() else None,
            "sub_status": r["sub_status"] or "",
            "started_at": r["started_at"] or "",
            "expires_at": expires_at or "",
            "payment_ref": r["payment_ref"] or "",
            "is_premium": is_premium,
        })

    return render_template(
        "admin/users.html",
        total_users=total_users,
        total_premium_active=total_premium_active,
        users=users,
        q=q,
        now_utc=now
    )


@app.route("/admin/users/<int:user_id>/adjust-expiry", methods=["POST"])
@login_required
def admin_adjust_user_expiry(user_id: int):
    """Admin: เพิ่ม/ลดวันหมดอายุ Premium ของผู้ใช้"""
    if not _is_admin():
        abort(403)

    delta_days_raw = (request.form.get("days") or "").strip()
    try:
        delta_days = int(delta_days_raw)
    except Exception:
        delta_days = 0

    # ถ้าไม่มี sub แล้ว admin อยากให้เริ่ม premium ใหม่: ใช้ grant_premium ได้
    if delta_days == 0:
        flash("จำนวนวันไม่ถูกต้อง", "error")
        return redirect(url_for("admin_users", q=request.args.get("q") or ""))

    try:
        # ปรับที่ subscription ล่าสุด (ถ้ามี) ไม่งั้นให้ grant premium ใหม่
        sub = UserSubscription.get_active_subscription(user_id)
        if sub:
            UserSubscription.adjust_expiry(sub["id"], delta_days)
            flash(f"ปรับวันหมดอายุ {delta_days:+d} วัน เรียบร้อย", "success")
        else:
            # ไม่มี active -> สร้างใหม่เป็น admin grant
            UserSubscription.grant_premium(user_id, max(delta_days, 1), reason="admin_adjust")
            flash(f"สร้าง Premium ใหม่ {max(delta_days,1)} วัน เรียบร้อย", "success")
    except Exception as e:
        flash(f"ปรับวันหมดอายุไม่สำเร็จ: {e}", "error")

    return redirect(url_for("admin_users", q=request.args.get("q") or ""))

# ==============================================================================
# Admin
# ==============================================================================
# """  # stray triple-quote kept as comment


if __name__ == "__main__":
    app.run(debug=os.environ.get("FLASK_ENV") == "development", host="0.0.0.0", port=int(os.environ.get("PORT", "5000")))
